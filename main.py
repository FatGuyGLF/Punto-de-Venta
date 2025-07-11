# --- Importaciones de bibliotecas estándar y de terceros ---
import tkinter as tk
from tkinter import messagebox, simpledialog, ttk, filedialog
import configparser
import os
import csv
import shutil
from datetime import datetime

# --- Importaciones de módulos locales ---
from database import Database
from models import Usuario, Producto, Categoria, Venta, Devolucion, Gasto

# --- Importaciones para funcionalidades específicas ---
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg 
from fpdf import FPDF
import openpyxl
import webbrowser

# --- Constantes Globales ---
CONFIG_FILE = 'config.info'

# --- Funciones Auxiliares ---

def generarTicketPdf(carrito, totalFinal, idVenta, pagoInfo):
    """
    Genera un archivo PDF con el formato de un ticket de compra.
    
    Args:
        carrito (list): Lista de diccionarios de los productos vendidos.
        totalFinal (float): Monto total de la venta.
        idVenta (int): ID de la venta para el encabezado del ticket.
        pagoInfo (dict): Información sobre el método de pago, monto recibido y cambio.
    
    Returns:
        str: El nombre del archivo PDF generado.
    Raises:
        Exception: Si ocurre un error durante la generación del PDF.
    """
    try:
        # Configuración inicial del documento PDF en un formato de 80mm de ancho (típico para impresoras de tickets)
        pdf = FPDF(orientation='P', unit='mm', format=(80, 200))
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=5) # Salto de página automático
        pdf.set_font("Courier", "", 10) # Fuente monoespaciada para alineación
        pdf.set_margins(5, 5, 5)

        # Encabezado del Ticket
        pdf.set_font("Courier", "B", 12)
        pdf.cell(0, 5, "Papeleria Flores", 0, 1, "C") # Título centrado
        pdf.cell(0, 5, "--- RECIBO DE COMPRA ---", 0, 1, "C")
        pdf.set_font("Courier", "", 8)
        pdf.cell(35, 5, f"Ticket No: {idVenta}", 0, 0, "L")
        pdf.cell(35, 5, datetime.now().strftime("%d/%m/%Y %H:%M"), 0, 1, "R")
        pdf.ln(3) # Salto de línea pequeño

        # Cabeceras de la tabla de productos
        pdf.set_font("Courier", "", 9)
        pdf.cell(40, 5, "Descripcion", 0, 0, "L")
        pdf.cell(10, 5, "Cant", 0, 0, "C")
        pdf.cell(20, 5, "Precio", 0, 1, "R")
        pdf.cell(0, 2, "-" * 38, 0, 1, "C") # Línea separadora

        # Contenido del carrito
        subtotal = sum(item['subtotal'] for item in carrito)
        for item in carrito:
            nombreProd = item['nombre'][:25] # Trunca el nombre del producto si es muy largo
            pdf.cell(40, 5, nombreProd, 0, 0, "L")
            pdf.cell(10, 5, str(item['cantidad']), 0, 0, "C")
            pdf.cell(20, 5, f"${item['subtotal']:.2f}", 0, 1, "R")
        
        pdf.cell(0, 3, "-" * 38, 0, 1, "C")
        pdf.ln(1)
        
        # Sección de Totales
        pdf.set_font("Courier", "", 9)
        pdf.cell(40, 5, "Subtotal:", 0, 0, "R")
        pdf.cell(30, 5, f"${subtotal:.2f}", 0, 1, "R")

        descuento = subtotal - totalFinal
        if descuento > 0.01: # Muestra el descuento solo si es significativo
            pdf.cell(40, 5, "Descuento:", 0, 0, "R")
            pdf.cell(30, 5, f"-${descuento:.2f}", 0, 1, "R")

        pdf.set_font("Courier", "B", 10)
        pdf.cell(40, 6, "Total:", 0, 0, "R")
        pdf.cell(30, 6, f"${totalFinal:.2f}", 0, 1, "R")
        
        # Información del Pago
        pdf.set_font("Courier", "", 9)
        pdf.cell(40, 6, "Metodo Pago:", 0, 0, "R")
        pdf.cell(30, 6, pagoInfo['metodo'], 0, 1, "R")
        
        if pagoInfo['metodo'] == 'Efectivo' and pagoInfo.get('efectivo', 0) > 0:
            pdf.cell(40, 6, "Recibido:", 0, 0, "R")
            pdf.cell(30, 6, f"${pagoInfo['efectivo']:.2f}", 0, 1, "R")
            pdf.cell(40, 6, "Cambio:", 0, 0, "R")
            pdf.cell(30, 6, f"${pagoInfo['cambio']:.2f}", 0, 1, "R")

        # Pie del Ticket
        pdf.ln(5)
        pdf.set_font("Courier", "B", 12)
        pdf.cell(0, 8, "¡GRACIAS POR SU COMPRA!", 0, 1, "C")
        
        nombreArchivo = f"Ticket Venta {idVenta}.pdf"
        pdf.output(nombreArchivo)
        return nombreArchivo
    except Exception as e:
        messagebox.showerror("Error de PDF", f"No se pudo generar el ticket:\n{e}")
        raise e

# --- Clases de la Interfaz Gráfica (GUI) ---

class LoginWindow(tk.Toplevel):
    """
    Ventana inicial de inicio de sesión para la aplicación.
    Es la primera ventana que ve el usuario.
    """
    def __init__(self, parent, onLoginSuccessCallback, db_instance):
        super().__init__(parent)
        self.rootApp = parent # La ventana raíz de Tk(), que está oculta
        self.onLoginSuccess = onLoginSuccessCallback # Función a llamar si el login es exitoso
        self.db = db_instance
        self.title("Iniciar Sesión")
        self.geometry("300x200")
        self.resizable(False, False)
        
        # --- Centrado de la ventana ---
        parent.update_idletasks() # Asegura que las dimensiones de la ventana padre estén actualizadas
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (self.winfo_width() // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")
        
        # --- Configuración de comportamiento de la ventana ---
        self.protocol("WM_DELETE_WINDOW", self.rootApp.destroy) # Cierra toda la app si se cierra esta ventana
        self.grab_set() # Hace que esta ventana sea modal, bloqueando la interacción con otras ventanas

        # --- Creación de Widgets ---
        frame = tk.Frame(self, padx=10, pady=10)
        frame.pack(expand=True)
        
        self.username = tk.StringVar(value=self.readRememberedUser())
        self.password = tk.StringVar()
        self.rememberMe = tk.BooleanVar(value=bool(self.readRememberedUser()))
        
        tk.Label(frame, text="Usuario:").grid(row=0, column=0, sticky="w", pady=2)
        tk.Entry(frame, textvariable=self.username).grid(row=0, column=1, pady=2)
        tk.Label(frame, text="Contraseña:").grid(row=1, column=0, sticky="w", pady=2)
        passwordEntry = tk.Entry(frame, textvariable=self.password, show="*")
        passwordEntry.grid(row=1, column=1, pady=2)
        tk.Checkbutton(frame, text="Recordar usuario", variable=self.rememberMe).grid(row=2, columnspan=2, pady=5)
        tk.Button(frame, text="Ingresar", command=self.login).grid(row=3, columnspan=2, pady=10)
        
        # --- Binds y Foco ---
        passwordEntry.bind("<Return>", lambda e: self.login()) # Permite ingresar con la tecla Enter
        passwordEntry.focus() # Pone el cursor en el campo de contraseña al abrir

    def login(self):
        """Intenta autenticar al usuario con las credenciales ingresadas."""
        user, pwd = self.username.get(), self.password.get()
        with self.db.connect() as conn:
            role = Usuario.verifyCredentials(conn, user, pwd)
        
        if role: # Si las credenciales son válidas
            self.handleRememberMe(user)
            self.destroy() # Cierra la ventana de login
            self.onLoginSuccess(role, user) # Llama al callback para abrir la siguiente ventana
        else:
            messagebox.showerror("Error", "Usuario o contraseña incorrectos.", parent=self)

    def readRememberedUser(self):
        """Lee el nombre de usuario del archivo de configuración si existe."""
        config = configparser.ConfigParser()
        if os.path.exists(CONFIG_FILE):
            config.read(CONFIG_FILE)
            return config.get('Login', 'username', fallback='')
        return ''

    def handleRememberMe(self, username):
        """Guarda o borra el nombre de usuario en el archivo de configuración según el estado del checkbox."""
        config = configparser.ConfigParser()
        # Si 'rememberMe' está activado, guarda el usuario; de lo contrario, guarda una cadena vacía.
        config['Login'] = {'username': username} if self.rememberMe.get() else {'username': ''}
        with open(CONFIG_FILE, 'w') as f:
            config.write(f)

class DashboardWindow(tk.Toplevel):
    """
    Panel de control principal para administradores.
    Muestra métricas clave, gráficos de ventas y proporciona navegación a otras ventanas administrativas.
    """
    # Paleta de colores para una apariencia consistente en la UI
    COLOR_PRINCIPAL = "#006400"
    COLOR_SECUNDARIO = "#2E8B57"
    COLOR_TERCIARIO = "#3CB371"
    COLOR_FONDO_GRAFICO = "#f0f0f0"
    COLOR_TEXTO_GRAFICO = "#333333"
    COLOR_GRID = "#dcdcdc"

    def __init__(self, root, username, db_instance):
        super().__init__(root)
        self.rootApp = root
        self.username = username
        self.db = db_instance
        self.title(f"Dashboard - Admin: {username}")
        self.geometry("1400x700")
        self.protocol("WM_DELETE_WINDOW", self.rootApp.destroy)
        self.configure(bg=self.COLOR_FONDO_GRAFICO)
        
        # Estilo para las pestañas del Notebook
        style = ttk.Style(self)
        style.configure("TNotebook.Tab", font=('Arial','11','bold'), padding=[10, 5])

        # --- Panel de Navegación Izquierdo ---
        left_nav_frame = tk.Frame(self, bg="#2c3e50", width=220)
        left_nav_frame.pack(side="left", fill="y")
        
        tk.Label(left_nav_frame, text=f"Bienvenido,\n{username}", font=("Arial", 16, "bold"), bg="#2c3e50", fg="white", wraplength=180).pack(pady=20, padx=10)
        self.createNavButton(left_nav_frame, "🛒", "Punto de Venta", self.openPos)
        self.createNavButton(left_nav_frame, "📦", "Inventario", lambda: self.openAdminWindow(AdminInventarioWindow))
        self.createNavButton(left_nav_frame, "👥", "Usuarios", lambda: self.openAdminWindow(AdminUsuariosWindow, self.username))
        self.createNavButton(left_nav_frame, "📈", "Finanzas", lambda: self.openAdminWindow(ReportesDevolucionesWindow))
        self.createNavButton(left_nav_frame, "🛠️", "Herramientas", lambda: self.openAdminWindow(HerramientasWindow))
        
        # --- Contenido Principal Derecho ---
        right_frame = tk.Frame(self, bg=self.COLOR_FONDO_GRAFICO)
        right_frame.pack(side="right", expand=True, fill="both")
        
        # Se usa un Notebook (pestañas) para organizar el contenido
        self.notebook = ttk.Notebook(right_frame)
        self.notebook.pack(expand=True, fill="both", padx=10, pady=10)

        # Creación de la Pestaña de Resumen
        self.tabResumen = tk.Frame(self.notebook, bg=self.COLOR_FONDO_GRAFICO)
        self.notebook.add(self.tabResumen, text='Resumen General')
        self.createResumenWidgets()

        # Creación de la Pestaña de Análisis
        self.tabAnalisis = tk.Frame(self.notebook, bg=self.COLOR_FONDO_GRAFICO)
        self.notebook.add(self.tabAnalisis, text='Análisis de Ventas')
        self.createAnalisisWidgets()

    def createNavButton(self, parent, icon, text, command):
        """Crea un botón estilizado para el panel de navegación izquierdo."""
        btnFrame = tk.Frame(parent, bg="#34495e", cursor="hand2")
        btnFrame.pack(fill="x", pady=2, padx=10)
        iconLabel = tk.Label(btnFrame, text=icon, font=("Arial", 18), bg="#34495e", fg="white", anchor="w")
        iconLabel.pack(side="left", padx=(15, 10), pady=10)
        textLabel = tk.Label(btnFrame, text=text, font=("Arial", 12), bg="#34495e", fg="white", anchor="w")
        textLabel.pack(side="left", fill="x", expand=True)
        # Binds para manejar eventos de clic y hover, cambiando el color de fondo
        for widget in [btnFrame, iconLabel, textLabel]:
            widget.bind("<Button-1>", lambda event: command()) # Ejecuta el comando al hacer clic
            widget.bind("<Enter>", lambda e, b=btnFrame: b.config(bg="#4e6a85")) # Cambia color al pasar el mouse
            widget.bind("<Leave>", lambda e, b=btnFrame: b.config(bg="#34495e")) # Restaura el color al quitar el mouse

    def createResumenWidgets(self):
        """Crea los widgets para la pestaña de Resumen General (métricas y gráfica principal)."""
        resumenFrame = tk.Frame(self.tabResumen, bg=self.COLOR_FONDO_GRAFICO, padx=10, pady=10)
        resumenFrame.pack(expand=True, fill="both")
        
        # Frame para las tarjetas de métricas
        metricsFrame = tk.Frame(resumenFrame, bg=self.COLOR_FONDO_GRAFICO)
        metricsFrame.pack(fill="x", pady=10)
        self.ventasVar, self.ticketsVar, self.stockVar = tk.StringVar(), tk.StringVar(), tk.StringVar()
        
        # Botones que actúan como "tarjetas" de métricas principales
        tk.Button(metricsFrame, textvariable=self.ventasVar, font=("Arial", 16, "bold"), relief="raised", bd=2, bg=self.COLOR_TERCIARIO, fg="white", width=16, height=3, wraplength=160, command=self.openVentasReporte).pack(side="left", expand=True, fill="both", padx=10)
        tk.Button(metricsFrame, textvariable=self.ticketsVar, font=("Arial", 16, "bold"), relief="raised", bd=2, bg="#3498DB", fg="white", width=16, height=3, command=self.openLibroDiario).pack(side="left", expand=True, fill="both", padx=10)
        tk.Button(metricsFrame, textvariable=self.stockVar, font=("Arial", 16, "bold"), relief="raised", bd=2, bg="#E74C3C", fg="white", width=16, height=3, command=self.openBajoStock).pack(side="left", expand=True, fill="both", padx=10)

        # Frame donde se dibujará la gráfica de ventas
        self.graficaVentasFrame = tk.Frame(resumenFrame, bg=self.COLOR_FONDO_GRAFICO)
        self.graficaVentasFrame.pack(expand=True, fill="both", pady=10)
        self.updateDashboardMetrics()

    def openBajoStock(self):
        """Abre la ventana que muestra productos con bajo stock."""
        self.openAdminWindow(LowStockWindow)

    def openVentasReporte(self):
        """Abre la ventana de finanzas directamente en el reporte de ventas del día."""
        config = {'initial_tab': 1, 'reporte': 'ventas', 'periodo': 'dia'}
        self.openAdminWindow(ReportesDevolucionesWindow, config)
    
    def openLibroDiario(self):
        """Abre la ventana de finanzas directamente en el libro diario del día."""
        config = {'initial_tab': 2, 'periodo': 'dia'}
        self.openAdminWindow(ReportesDevolucionesWindow, config)

    def createAnalisisWidgets(self):
        """Crea los widgets para la pestaña de Análisis de Ventas (gráficos por categoría y top productos)."""
        analisisFrame = tk.Frame(self.tabAnalisis, bg=self.COLOR_FONDO_GRAFICO, padx=10, pady=10)
        analisisFrame.pack(expand=True, fill="both")
        
        # Controles para cambiar el período de tiempo de los gráficos
        controlesFrame = tk.Frame(analisisFrame, bg=self.COLOR_FONDO_GRAFICO)
        controlesFrame.pack(fill="x", pady=5)
        tk.Label(controlesFrame, text="Ver período:", font=("Arial", 11), bg=self.COLOR_FONDO_GRAFICO).pack(side="left", padx=(0,10))
        self.periodoAnalisis = tk.StringVar(value="semana")
        ttk.Radiobutton(controlesFrame, text="Día", variable=self.periodoAnalisis, value="dia", command=self.updateAnalisisGraphs).pack(side="left")
        ttk.Radiobutton(controlesFrame, text="Semana", variable=self.periodoAnalisis, value="semana", command=self.updateAnalisisGraphs).pack(side="left")
        ttk.Radiobutton(controlesFrame, text="Mes", variable=self.periodoAnalisis, value="mes", command=self.updateAnalisisGraphs).pack(side="left")
        
        # Contenedor para los dos gráficos de análisis
        graficasContainer = tk.Frame(analisisFrame, bg=self.COLOR_FONDO_GRAFICO)
        graficasContainer.pack(expand=True, fill="both", pady=10)
        
        self.graficaCategoriasFrame = tk.LabelFrame(graficasContainer, text=" Ingresos por Categoría ", font=("Arial", 11), bg=self.COLOR_FONDO_GRAFICO, fg=self.COLOR_TEXTO_GRAFICO, bd=1)
        self.graficaCategoriasFrame.pack(side="left", expand=True, fill="both", padx=(0, 10), ipady=5)
        
        self.graficaTopProductosFrame = tk.LabelFrame(graficasContainer, text=" Top 5 Productos por Ingresos ", font=("Arial", 11), bg=self.COLOR_FONDO_GRAFICO, fg=self.COLOR_TEXTO_GRAFICO, bd=1)
        self.graficaTopProductosFrame.pack(side="left", expand=True, fill="both", padx=(10, 0), ipady=5)

        self.updateAnalisisGraphs()

    def updateDashboardMetrics(self):
        """Actualiza los valores de las tarjetas de métricas y la gráfica de ventas diarias."""
        try:
            with self.db.connect() as conn:
                data = Venta.getDashboardData(conn)
            self.ventasVar.set(f"Ventas Hoy\n${data['ventasNetasHoy']:.2f}")
            self.ticketsVar.set(f"Tickets Hoy\n{data['numTicketsHoy']}")
            self.stockVar.set(f"Bajo Stock\n{data['productosBajoStock']} items")
            self.createDailySalesGraph(self.graficaVentasFrame)
        except Exception as e:
            messagebox.showerror("Error de Dashboard", f"No se pudieron cargar los datos: {e}")
    
    def clearFrame(self, frame):
        """Elimina todos los widgets dentro de un frame, útil para refrescar gráficos."""
        for widget in frame.winfo_children():
            widget.destroy()

    def createDailySalesGraph(self, parent):
        """Crea y muestra la gráfica de barras de ventas de los últimos 7 días."""
        self.clearFrame(parent)
        try:
            with self.db.connect() as conn:
                datos = Venta.getVentasUltimosDias(conn, dias=7)
            
            dias_semana = list(datos.keys())
            valores = list(datos.values())

            # Creación de la figura y el eje de Matplotlib
            fig = Figure(figsize=(8, 4), dpi=100, facecolor=self.COLOR_FONDO_GRAFICO)
            ax = fig.add_subplot(111)
            ax.set_facecolor(self.COLOR_FONDO_GRAFICO)

            # Dibujo de las barras
            bars = ax.bar(dias_semana, valores, color=self.COLOR_SECUNDARIO, width=0.6, zorder=2)
            
            # Estilo de la gráfica
            ax.set_title("Ventas de los Últimos 7 Días", fontsize=14, color=self.COLOR_TEXTO_GRAFICO, pad=20)
            ax.set_ylabel("Ventas ($)", fontsize=10, color=self.COLOR_TEXTO_GRAFICO)
            ax.tick_params(axis='x', colors=self.COLOR_TEXTO_GRAFICO)
            ax.tick_params(axis='y', colors=self.COLOR_TEXTO_GRAFICO)
            ax.grid(axis='y', linestyle='--', color=self.COLOR_GRID, zorder=1)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.spines['left'].set_color(self.COLOR_GRID)
            ax.spines['bottom'].set_color(self.COLOR_GRID)
            fig.tight_layout() # Ajusta el layout para que no se corten las etiquetas

            # --- Funcionalidad de Hover para mostrar valores ---
            # Se crea una anotación (texto flotante) que estará oculta por defecto
            annot = ax.annotate("", xy=(0,0), xytext=(0,15), textcoords="offset points",
                                bbox=dict(boxstyle="round,pad=0.4", fc=self.COLOR_PRINCIPAL, ec="none", alpha=0.9),
                                arrowprops=dict(arrowstyle="->", connectionstyle="arc3", color=self.COLOR_PRINCIPAL),
                                ha="center", color="white", fontweight="bold", visible=False)
            
            def on_hover(event):
                """Función que se ejecuta cada vez que el mouse se mueve sobre la gráfica."""
                vis = annot.get_visible()
                if event.inaxes == ax: # Si el evento ocurrió dentro del eje de la gráfica
                    for i, bar in enumerate(bars):
                        if bar.contains(event)[0]: # Si el mouse está sobre una de las barras
                            x, y = bar.get_x() + bar.get_width() / 2, bar.get_height()
                            annot.xy = (x, y) # Posiciona la anotación sobre la barra
                            annot.set_text(f" ${valores[i]:.2f} ") # Pone el valor de la barra en el texto
                            annot.set_visible(True) # La hace visible
                            fig.canvas.draw_idle() # Redibuja
                            return
                if vis: # Si la anotación estaba visible pero el mouse ya no está sobre ninguna barra
                    annot.set_visible(False)
                    fig.canvas.draw_idle()

            # Incrustar la figura de Matplotlib en Tkinter
            canvas = FigureCanvasTkAgg(fig, master=parent)
            canvas.draw()
            canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
            # Conectar el evento de movimiento del mouse a la función on_hover
            canvas.mpl_connect("motion_notify_event", on_hover)

        except Exception as e:
            tk.Label(parent, text=f"Error al generar gráfica:\n{e}", bg=self.COLOR_FONDO_GRAFICO).pack(expand=True)
    
    def updateAnalisisGraphs(self):
        """Actualiza las gráficas de la pestaña de análisis según el período seleccionado."""
        periodo = self.periodoAnalisis.get()
        self.createCategorySalesGraph(self.graficaCategoriasFrame, periodo)
        self.createTopProductsGraph(self.graficaTopProductosFrame, periodo)

    def createCategorySalesGraph(self, parent, periodo):
        """Crea y muestra la gráfica de dona de ingresos por categoría."""
        self.clearFrame(parent)
        try:
            with self.db.connect() as conn:
                datos = Venta.getVentasPorCategoria(conn, periodo)
            if not datos:
                tk.Label(parent, text=f"No hay datos de ventas para este período.", bg=self.COLOR_FONDO_GRAFICO).pack(expand=True, fill="both", pady=20)
                return
            
            nombres, valores = [d[0] for d in datos], [d[1] for d in datos]
            
            fig = Figure(figsize=(5.5, 5), dpi=100, facecolor=self.COLOR_FONDO_GRAFICO)
            fig.subplots_adjust(bottom=0.3) # Aumentar el margen inferior para la leyenda
            ax = fig.add_subplot(111)
            
            colores = [self.COLOR_PRINCIPAL, self.COLOR_SECUNDARIO, self.COLOR_TERCIARIO, "#98FB98", "#20B2AA", "#008080"]
            
            # Creación de la gráfica de pie (dona)
            wedges, texts, autotexts = ax.pie(valores, labels=None, 
                                             autopct=lambda pct: f'{pct:.1f}%' if pct > 5 else '', # Muestra porcentaje si es > 5%
                                             wedgeprops=dict(width=0.5, ec=self.COLOR_FONDO_GRAFICO, lw=3), # El `width` crea el efecto de dona
                                             startangle=90, colors=colores, pctdistance=0.75)
            
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
                autotext.set_fontsize(9)

            # Texto central en la grafica de dona
            total = sum(valores)
            ax.text(0, 0, f'Total\n${total:,.2f}', ha='center', va='center', fontsize=16, color=self.COLOR_TEXTO_GRAFICO, fontweight='bold')
            
            # Leyenda de la gráfica
            ax.legend(wedges, nombres, 
                      title="Categorías",
                      loc='upper center', 
                      bbox_to_anchor=(0.5, -0.05), # Posiciona la leyenda debajo del gráfico
                      ncol=min(len(nombres), 3),   # Máximo 3 columnas para que no sea muy ancha
                      fontsize=9,
                      frameon=False)
            
            canvas = FigureCanvasTkAgg(fig, master=parent)
            canvas.draw()
            canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        except Exception as e:
            tk.Label(parent, text=f"Error al generar gráfica:\n{e}", bg=self.COLOR_FONDO_GRAFICO).pack(expand=True)

    def createTopProductsGraph(self, parent, periodo):
        """Crea y muestra la gráfica de barras horizontales de los 5 productos más vendidos por ingresos."""
        self.clearFrame(parent)
        try:
            with self.db.connect() as conn:
                datos = Venta.getTopProductos(conn, periodo)
            if not datos:
                tk.Label(parent, text=f"No hay productos vendidos en este período.", bg=self.COLOR_FONDO_GRAFICO).pack(expand=True, fill="both", pady=20)
                return
            
            nombres = [d[0] for d in datos]
            valores = [d[1] for d in datos]
            
            fig = Figure(figsize=(5.5, 5), dpi=100, facecolor=self.COLOR_FONDO_GRAFICO)
            # Usar ajuste manual con margen izquierdo grande para que los nombres de producto no se corten
            fig.subplots_adjust(left=0.45, right=0.95, top=0.9, bottom=0.15)
            ax = fig.add_subplot(111)
            ax.set_facecolor(self.COLOR_FONDO_GRAFICO)

            # Gráfica de barras horizontales
            bars = ax.barh(nombres, valores, color=self.COLOR_TERCIARIO, height=0.6, zorder=2)
            ax.invert_yaxis() # El producto con más ingresos aparece arriba
            
            # Estilo de la gráfica
            ax.set_xlabel("Ingresos ($)", fontsize=10, color=self.COLOR_TEXTO_GRAFICO)
            ax.tick_params(axis='x', colors=self.COLOR_TEXTO_GRAFICO)
            ax.tick_params(axis='y', colors=self.COLOR_TEXTO_GRAFICO, labelsize=9)
            ax.grid(axis='x', linestyle='--', color=self.COLOR_GRID, zorder=1)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.spines['bottom'].set_color(self.COLOR_GRID)
            ax.spines['left'].set_visible(False)
            ax.tick_params(axis='y', length=0) # Oculta las pequeñas marcas junto a las etiquetas Y
            
            # Funcionalidad de Hover (similar a la gráfica de ventas diarias)
            annot = ax.annotate("", xy=(0,0), xytext=(15,0), textcoords="offset points",
                                bbox=dict(boxstyle="round,pad=0.4", fc=self.COLOR_PRINCIPAL, ec="none", alpha=0.9),
                                arrowprops=dict(arrowstyle="->", connectionstyle="arc3", color=self.COLOR_PRINCIPAL),
                                ha="left", va="center", color="white", fontweight="bold", visible=False)

            def on_hover(event):
                vis = annot.get_visible()
                if event.inaxes == ax:
                    for i, bar in enumerate(bars):
                        if bar.contains(event)[0]:
                            x, y = bar.get_width(), bar.get_y() + bar.get_height() / 2
                            annot.xy = (x, y)
                            annot.set_text(f" ${valores[i]:.2f} ")
                            annot.set_visible(True)
                            fig.canvas.draw_idle()
                            return
                if vis:
                    annot.set_visible(False)
                    fig.canvas.draw_idle()

            canvas = FigureCanvasTkAgg(fig, master=parent)
            canvas.draw()
            canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
            canvas.mpl_connect("motion_notify_event", on_hover)

        except Exception as e:
            tk.Label(parent, text=f"Error al generar gráfica:\n{e}", bg=self.COLOR_FONDO_GRAFICO).pack(expand=True)

    def openPos(self):
        """Abre la ventana del Punto de Venta y oculta el dashboard."""
        self.withdraw() # Oculta la ventana actual (Dashboard)
        PuntoVentaApp(self, 'admin', self.username, self.db)
    
    def openAdminWindow(self, windowClass, *args):
        """
        Abre una ventana administrativa de forma modal.
        Al cerrarla, refresca las métricas y gráficos del dashboard.
        
        Args:
            windowClass: La clase de la ventana a abrir (ej. AdminInventarioWindow).
            *args: Argumentos adicionales para el constructor de la clase.
        """
        adminWindow = windowClass(self, self.db, *args)
        adminWindow.grab_set() # Bloquea la interacción con el Dashboard
        self.wait_window(adminWindow) # Espera hasta que la ventana administrativa sea cerrada
        try:
            # Una vez cerrada la ventana modal, actualiza los datos del Dashboard
            self.updateDashboardMetrics()
            self.updateAnalisisGraphs()
        except tk.TclError:
            # Este error puede ocurrir si la ventana principal ya fue destruida, se ignora.
            pass

class PuntoVentaApp(tk.Toplevel):
    """
    La interfaz principal para realizar ventas (cajeros y administradores).
    Permite buscar productos, añadirlos al carrito, aplicar descuentos y procesar el pago.
    """
    def __init__(self, parent, userRole, username, db_instance):
        super().__init__(parent)
        self.parent = parent
        self.userRole = userRole
        self.username = username
        self.db = db_instance
        self.title(f"Punto de Venta (Usuario: {self.username})")
        self.geometry("1200x600")
        self.protocol("WM_DELETE_WINDOW", self.onClose)

        # --- Estructuras de datos principales ---
        self.carrito = [] # Lista para almacenar los productos de la venta actual
        self.descuentoPorcentaje = 0.0

        # --- Creación de Widgets ---
        mainFrame = tk.Frame(self, padx=10, pady=10)
        mainFrame.pack(fill=tk.BOTH, expand=True)
        
        # Frame de búsqueda de productos
        searchFrame = tk.Frame(mainFrame)
        searchFrame.pack(fill=tk.X, pady=5)
        tk.Label(searchFrame, text="Buscar producto:", font=("Arial", 14)).pack(side=tk.LEFT)
        self.searchVar = tk.StringVar()
        self.searchVar.trace_add("write", self.onSearchEntryChange) # Llama a la función cada vez que el texto cambia
        self.entryCodigo = tk.Entry(searchFrame, font=("Arial", 14), textvariable=self.searchVar)
        self.entryCodigo.pack(fill=tk.X, expand=True, side=tk.LEFT, padx=5)
        self.entryCodigo.bind("<Return>", self.onEnterInSearch) # Evento para la tecla Enter
        self.entryCodigo.bind("<Down>", self.focusOnSuggestions) # Evento para la flecha abajo
        self.entryCodigo.focus()

        # Listbox para mostrar sugerencias de búsqueda (inicialmente oculta)
        self.suggestionListbox = tk.Listbox(self, font=("Courier", 11))
        self.suggestionListbox.bind("<Double-Button-1>", self.onSuggestionSelect)
        self.suggestionListbox.bind("<Return>", self.onSuggestionSelect)
        self.searchResults = [] # Almacena los resultados de la búsqueda actual

        # Frame para mostrar el carrito de compras
        carritoFrame = tk.LabelFrame(mainFrame, text="Carrito", padx=10, pady=10)
        carritoFrame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.listaCarrito = tk.Listbox(carritoFrame, font=("Courier", 12)) # Fuente monoespaciada para alinear texto
        self.listaCarrito.pack(fill=tk.BOTH, expand=True)

        # Frame para los botones de acción del carrito
        botonesFrame = tk.Frame(mainFrame)
        botonesFrame.pack(fill=tk.X, pady=5)
        tk.Button(botonesFrame, text="Eliminar", command=self.deleteProduct, bg="#E74C3C", fg="white").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
        tk.Button(botonesFrame, text="Modificar", command=self.modifyProduct, bg="#F1C40F").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
        tk.Button(botonesFrame, text="Descuento", command=self.applyDiscount, bg="#E67E22", fg="white").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
        
        # Botón de navegación (diferente para admin y cajero)
        if self.userRole == 'admin':
             tk.Button(botonesFrame, text="< Volver al Dashboard", command=self.onClose).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
        else:
            tk.Button(botonesFrame, text="Cerrar Sesión", command=self.onClose).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        # Frame para el botón de confirmar venta
        ventaFrame = tk.Frame(mainFrame)
        ventaFrame.pack(fill=tk.X, pady=5)
        tk.Button(ventaFrame, text="Confirmar Venta", command=self.confirmSale, font=("Arial", 14, "bold"), bg="#2ECC71", fg="white").pack(expand=True, fill=tk.X)
        
        self.updateCartList() # Actualiza la lista del carrito para mostrar los totales iniciales

    def openSweetsDialog(self):
        """Abre un diálogo especial para la venta rápida de dulces."""
        # Busca el ID de la categoría 'dulces'
        with self.db.connect() as conn:
            categorias = {cat[1].lower(): cat[0] for cat in Categoria.getAll(conn)}
        
        id_categoria_dulces = categorias.get('dulces')
        if not id_categoria_dulces: return # No hace nada si la categoría no existe
            
        dialog = DialogoVentaDulces(self, self.db, id_categoria_dulces)
        self.wait_window(dialog)

        # Si se seleccionaron dulces, los añade al carrito principal
        if dialog.dulcesSeleccionados:
            for producto, cantidad in dialog.dulcesSeleccionados.values():
                self.addProductToCart(producto, cantidad=cantidad, check_category=False)
    
    def onSearchEntryChange(self, *args):
        """Se activa cada vez que el usuario escribe en el campo de búsqueda."""
        userInput = self.searchVar.get()
        # Solo busca si el input tiene al menos 2 caracteres (si es texto) o si es un número (código de barras)
        if len(userInput) < 2 and not userInput.isdigit():
            self.hideSuggestions()
            return
        
        with self.db.connect() as conn:
            if userInput.isdigit(): # Búsqueda por código de barras
                product_by_barcode = Producto.getByBarcode(conn, userInput)
                self.searchResults = [product_by_barcode] if product_by_barcode else []
            else: # Búsqueda por nombre
                self.searchResults = Producto.searchByName(conn, userInput)
        
        if self.searchResults:
            self.showSuggestions()
        else:
            self.hideSuggestions()

    def showSuggestions(self):
        """Muestra y posiciona la lista de sugerencias debajo del campo de búsqueda."""
        self.suggestionListbox.delete(0, tk.END)
        for product in self.searchResults:
            displayText = f"{product['nombre']} ({product['categoriaNombre'] or 'Sin Cat.'}) - Stock: {product['stock']}"
            self.suggestionListbox.insert(tk.END, displayText)
        
        # Calcula la posición absoluta del campo de búsqueda y la convierte a relativa a la ventana
        x_abs = self.entryCodigo.winfo_rootx()
        y_abs = self.entryCodigo.winfo_rooty() + self.entryCodigo.winfo_height()
        x_rel = x_abs - self.winfo_rootx()
        y_rel = y_abs - self.winfo_rooty()
        width = self.entryCodigo.winfo_width()
        
        # Posiciona el listbox de sugerencias
        self.suggestionListbox.place(x=x_rel, y=y_rel, width=width, height=150)

    def hideSuggestions(self):
        """Oculta la lista de sugerencias."""
        self.suggestionListbox.place_forget()

    def focusOnSuggestions(self, event):
        """Mueve el foco del campo de búsqueda a la lista de sugerencias al presionar la flecha abajo."""
        if self.suggestionListbox.winfo_viewable():
            self.suggestionListbox.focus_set()
            self.suggestionListbox.selection_set(0) # Selecciona el primer elemento

    def onSuggestionSelect(self, event):
        """Se activa al seleccionar un item de la lista de sugerencias (con Enter o doble clic)."""
        selection = self.suggestionListbox.curselection()
        if not selection: return
        
        selectedIndex = selection[0]
        producto = self.searchResults[selectedIndex]
        
        self.searchVar.set("") # Limpia el campo de búsqueda
        self.hideSuggestions()

        # Lógica especial para la categoría 'dulces'
        if producto.get('categoriaNombre', '').lower() == 'dulces':
            self.openSweetsDialog()
        else:
            self.addProductToCart(producto)
        self.entryCodigo.focus() # Devuelve el foco al campo de búsqueda

    def onEnterInSearch(self, event):
        """Maneja la pulsación de Enter en el campo de búsqueda."""
        # Si la lista de sugerencias está visible, Enter selecciona el primer item
        if self.suggestionListbox.winfo_viewable() and self.suggestionListbox.size() > 0:
            self.onSuggestionSelect(None)
            return

        # Si no hay sugerencias, intenta buscar por código de barras
        userInput = self.searchVar.get()
        if userInput.isdigit():
            with self.db.connect() as conn:
                producto = Producto.getByBarcode(conn, userInput)
            if producto:
                self.searchVar.set("")
                # Lógica especial para dulces o recargas
                if producto.get('categoriaNombre', '').lower() == 'dulces':
                    self.openSweetsDialog()
                else:
                    self.addProductToCart(producto)
        else:
             messagebox.showerror("Error", f"Producto no encontrado.", parent=self)

    def addProductToCart(self, producto, cantidad=1, check_category=True):
        """
        Añade un producto al carrito, manejando casos especiales y verificando el stock.
        
        Args:
            producto (dict): El diccionario del producto a añadir.
            cantidad (int): La cantidad a añadir.
            check_category (bool): Si es True, verifica si es de la categoría 'dulces'.
        """
        if check_category and producto.get('categoriaNombre', '').lower() == 'dulces':
            self.openSweetsDialog()
            return

        try:
            # --- Lógica para Recargas de Celular ---
            if producto["nombre"] == "Recarga Celular":
                recharge_amount = self.askRechargeAmount()
                if recharge_amount is None: return # El usuario canceló

                # El precio de venta de una recarga es el monto + $1 de comisión
                precio_final_recarga = float(recharge_amount) + 1.00
                # Busca si ya hay una recarga del mismo monto en el carrito
                itemEnCarrito = next((item for item in self.carrito if item["id"] == producto["idProducto"] and item["precio"] == precio_final_recarga), None)

                if itemEnCarrito: # Si ya existe, solo incrementa la cantidad
                    itemEnCarrito["cantidad"] += 1
                    itemEnCarrito["subtotal"] = itemEnCarrito["cantidad"] * itemEnCarrito["precio"]
                else: # Si no existe, la añade como un nuevo item
                    self.carrito.append({
                        "id": producto["idProducto"], "nombre": f"Recarga Celular ${recharge_amount:.2f}",
                        "precio": precio_final_recarga, "cantidad": 1, "subtotal": precio_final_recarga
                    })
            else:
                # --- Lógica para productos normales ---
                with self.db.connect() as conn:
                    currentStock = Producto.getById(conn, producto["idProducto"])["stock"]
                
                itemEnCarrito = next((item for item in self.carrito if item["id"] == producto["idProducto"]), None)
                cantidadEnCarrito = itemEnCarrito['cantidad'] if itemEnCarrito else 0

                # Verifica si hay suficiente stock
                if (cantidadEnCarrito + cantidad) > currentStock:
                    raise ValueError(f"No hay suficiente stock para '{producto['nombre']}'. Disponible: {currentStock}")

                if itemEnCarrito: # Si ya está en el carrito, actualiza la cantidad
                    itemEnCarrito["cantidad"] += cantidad
                    itemEnCarrito["subtotal"] = itemEnCarrito["cantidad"] * itemEnCarrito["precio"]
                else: # Si es nuevo, lo añade al carrito
                    self.carrito.append({
                        "id": producto["idProducto"], "nombre": producto["nombre"], 
                        "precio": producto["precioVenta"], "cantidad": cantidad, 
                        "subtotal": producto["precioVenta"] * cantidad
                    })
            
            self.updateCartList() # Refresca la vista del carrito

        except ValueError as e:
            messagebox.showerror("Stock insuficiente", str(e), parent=self)
        except Exception as e:
            messagebox.showerror("Error al añadir producto", f"Ocurrió un error: {e}", parent=self)

    def askRechargeAmount(self):
        """Muestra un diálogo para seleccionar el monto de una recarga."""
        recharge_options = [10, 15, 20, 30, 40, 50, 60, 70, 80, 90, 100, 150, 200, 300, 400, 500]
        dialog = tk.Toplevel(self)
        dialog.title("Seleccionar Monto de Recarga")
        dialog.transient(self); dialog.grab_set(); dialog.resizable(False, False)
        result_amount = None
        tk.Label(dialog, text="Seleccione el monto de la recarga:").pack(pady=10)
        amount_var = tk.DoubleVar()
        
        # Crea una grilla de Radiobuttons para los montos
        radio_frame = tk.Frame(dialog)
        radio_frame.pack(pady=5, padx=10)
        num_cols = 4
        for i, amount in enumerate(recharge_options):
            tk.Radiobutton(radio_frame, text=f"${amount:.2f}", variable=amount_var, value=amount).grid(row=i // num_cols, column=i % num_cols, padx=5, pady=2, sticky="w")
        if recharge_options: amount_var.set(recharge_options[0])

        def on_confirm():
            nonlocal result_amount
            result_amount = amount_var.get()
            dialog.destroy()

        tk.Button(dialog, text="Confirmar", command=on_confirm).pack(side=tk.LEFT, padx=10, pady=10)
        tk.Button(dialog, text="Cancelar", command=dialog.destroy).pack(side=tk.RIGHT, padx=10, pady=10)
        self.wait_window(dialog)
        return result_amount

    # Dentro de la clase PuntoVentaApp
    def onClose(self):
        try:
            if hasattr(self.parent, 'updateDashboardMetrics'):
                # Si es admin, solo muestra el dashboard y cierra esta ventana
                self.parent.deiconify()
                self.destroy()
            else:
                # Si es cajero, destruye la aplicación entera y termina.
                # No se necesita llamar a self.destroy() después, porque la app entera desaparece.
                self.parent.destroy()
        except tk.TclError:
            # Si la ventana ya fue destruida por otra acción, ignora el error.
            pass

    def updateCartList(self):
        """Borra y re-dibuja la lista del carrito con los datos actualizados."""
        self.listaCarrito.delete(0, tk.END)
        subtotal = sum(item['subtotal'] for item in self.carrito)
        descuentoMonto = subtotal * (self.descuentoPorcentaje / 100)
        totalFinal = subtotal - descuentoMonto
        
        # Inserta cada item del carrito
        for item in self.carrito:
            texto = f"{item['nombre']:<30} | Cant: {item['cantidad']:<3} | Subtotal: ${item['subtotal']:>8.2f}"
            self.listaCarrito.insert(tk.END, texto)
            
        # Inserta las líneas de totales
        self.listaCarrito.insert(tk.END, "")
        self.listaCarrito.insert(tk.END, f"{'Subtotal:':<43} ${subtotal:>8.2f}")
        if self.descuentoPorcentaje > 0:
            self.listaCarrito.insert(tk.END, f"{f'Descuento ({self.descuentoPorcentaje:.1f}%):':<43} -${descuentoMonto:>7.2f}")
            self.listaCarrito.insert(tk.END, "-"*53)
        self.listaCarrito.insert(tk.END, f"{'TOTAL A PAGAR:':<43} ${totalFinal:>8.2f}")

    def applyDiscount(self):
        """Abre un diálogo para aplicar un descuento porcentual a toda la venta."""
        porcentaje = simpledialog.askfloat("Aplicar Descuento", "Ingrese el porcentaje de descuento (%):", minvalue=0.0, maxvalue=100.0, parent=self)
        if porcentaje is not None:
            self.descuentoPorcentaje = porcentaje
            self.updateCartList()

    def confirmSale(self):
        """Inicia el proceso final de la venta."""
        if not self.carrito: 
            messagebox.showwarning("Advertencia", "El carrito está vacío.", parent=self)
            return
            
        # Calcula el total final
        subtotal = sum(item['subtotal'] for item in self.carrito)
        descuentoMonto = subtotal * (self.descuentoPorcentaje / 100)
        totalFinal = subtotal - descuentoMonto
        
        # Abre el diálogo de pago
        pagoDialog = DialogoPago(self, totalFinal)
        self.wait_window(pagoDialog)
        
        # Si el pago se confirmó en el diálogo
        if pagoDialog.resultado:
            pagoInfo = pagoDialog.resultado
            if messagebox.askyesno("Confirmar Venta", f"Total (con descuento): ${totalFinal:.2f}\n¿Proceder?", parent=self):
                try:
                    # Registra la venta en la base de datos
                    with self.db.connect() as conn:
                        ventaId = Venta.create(conn, self.carrito, pagoInfo['metodo'], descuentoMonto)
                    
                    # Genera el ticket en PDF
                    ticketFile = generarTicketPdf(self.carrito, totalFinal, ventaId, pagoInfo)
                    messagebox.showinfo("Venta Confirmada", f"Venta #{ventaId} completada.\nTicket generado: {ticketFile}", parent=self)
                    
                    # Reinicia el estado del POS para una nueva venta
                    self.descuentoPorcentaje = 0.0
                    self.carrito.clear()
                    self.updateCartList()
                except Exception as e:
                    messagebox.showerror("Error Crítico", f"Ocurrió un error al registrar la venta:\n{e}", parent=self)
    
    def deleteProduct(self):
        """Elimina el producto seleccionado del carrito."""
        try:
            if not self.listaCarrito.curselection(): return # No hay nada seleccionado
            index = self.listaCarrito.curselection()[0]
            # Asegurarse de que no se intente borrar una línea de total
            if index < len(self.carrito):
                del self.carrito[index]
            self.updateCartList()
        except IndexError: pass # Ignora errores si el índice es inválido
    
    def modifyProduct(self):
        """Modifica la cantidad de un producto ya existente en el carrito."""
        try:
            if not self.listaCarrito.curselection(): return
            index = self.listaCarrito.curselection()[0]
            if index >= len(self.carrito): return # No se puede modificar una línea de total
            
            item = self.carrito[index]

            # Las recargas no se pueden modificar en cantidad
            if item['nombre'].startswith("Recarga Celular"):
                messagebox.showinfo("Información", "Las recargas no se pueden modificar.", parent=self)
                return

            nuevaCantidad = simpledialog.askinteger("Modificar cantidad", f"Nueva cantidad para {item['nombre']}:", minvalue=0, parent=self)
            
            if nuevaCantidad is not None:
                # Se vuelve a checar el stock disponible
                with self.db.connect() as conn:
                    productoInfo = Producto.getById(conn, item['id'])
                if nuevaCantidad > productoInfo['stock']:
                    messagebox.showerror("Error", f"No hay suficiente stock. Disponible: {productoInfo['stock']}", parent=self)
                    return
                
                if nuevaCantidad == 0: # Si la nueva cantidad es 0, se elimina el item
                    self.carrito.remove(item)
                else: # Se actualiza la cantidad y el subtotal
                    item['cantidad'] = nuevaCantidad
                    item['subtotal'] = item['cantidad'] * item['precio']
                self.updateCartList()
        except IndexError: pass

class DialogoPago(tk.Toplevel):
    """
    Un diálogo modal para seleccionar el método de pago y procesar el cobro.
    Calcula el cambio si el pago es en efectivo.
    """
    def __init__(self, parent, total):
        super().__init__(parent)
        self.title("Método de Pago")
        self.total = total
        self.resultado = None # Almacenará el resultado del pago
        self.geometry("350x250")
        self.resizable(False, False)
        self.grab_set() # Hace la ventana modal
        self.protocol("WM_DELETE_WINDOW", self.destroy)

        # --- Creación de Widgets ---
        tk.Label(self, text=f"Total a Pagar: ${self.total:.2f}", font=("Arial", 16)).pack(pady=10)
        
        self.metodoPago = tk.StringVar(value="Efectivo")
        tk.Radiobutton(self, text="Efectivo", variable=self.metodoPago, value="Efectivo", command=self.toggleEfectivo).pack(anchor="w", padx=20)
        tk.Radiobutton(self, text="Tarjeta", variable=self.metodoPago, value="Tarjeta", command=self.toggleEfectivo).pack(anchor="w", padx=20)
        
        # Frame para el campo de efectivo (se muestra/oculta)
        self.efectivoFrame = tk.Frame(self)
        self.efectivoFrame.pack(pady=5)
        tk.Label(self.efectivoFrame, text="Efectivo Recibido:").grid(row=0, column=0, padx=5)
        self.entryEfectivo = tk.Entry(self.efectivoFrame)
        self.entryEfectivo.grid(row=0, column=1, padx=5)
        self.entryEfectivo.focus()
        
        tk.Button(self, text="Confirmar Pago", command=self.confirmar).pack(pady=20)
        self.bind("<Return>", lambda event: self.confirmar())

    def toggleEfectivo(self):
        """Activa o desactiva el campo de texto para el efectivo recibido."""
        if self.metodoPago.get() == "Efectivo":
            self.entryEfectivo.config(state='normal')
        else:
            self.entryEfectivo.config(state='disabled')

    def confirmar(self):
        """Valida el pago y cierra el diálogo."""
        metodo = self.metodoPago.get()
        if metodo == 'Efectivo':
            try:
                efectivoRecibido = float(self.entryEfectivo.get())
                if efectivoRecibido < self.total:
                    messagebox.showerror("Error", "El efectivo recibido no puede ser menor que el total.", parent=self)
                    return
                # Guarda la información del pago en el diccionario de resultado
                self.resultado = {"metodo": "Efectivo", "efectivo": efectivoRecibido, "cambio": efectivoRecibido - self.total}
            except (ValueError, TypeError):
                messagebox.showerror("Error", "Por favor, ingrese un monto válido.", parent=self)
                return
        else: # Si el pago es con Tarjeta
            self.resultado = {"metodo": "Tarjeta"}
        
        self.destroy() # Cierra la ventana de diálogo
        
class ReportesDevolucionesWindow(tk.Toplevel):
    """
    Ventana para la gestión financiera. Incluye:
    - Estado Financiero: Calcula el balance de caja.
    - Reportes de Ventas y Ganancias.
    - Libro Diario: Un historial de todas las transacciones.
    - Registro de Gastos.
    - Interfaz para iniciar Devoluciones.
    """
    def __init__(self, parent, db_instance, config={}, *args):
        super().__init__(parent)
        self.db = db_instance
        self.parent_dashboard = parent 
        self.title("Finanzas y Devoluciones")
        self.geometry("850x650")
        self.protocol("WM_DELETE_WINDOW", self.onClose)

        style = ttk.Style(self)
        style.configure("TNotebook.Tab", font=('Arial','11'), padding=[10, 5])
        
        # El Notebook es el componente principal que contiene las pestañas
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(expand=True, fill='both', padx=10, pady=10)

        # Creación de los frames para cada pestaña
        estadoFinancieroFrame = tk.Frame(self.notebook)
        reportesFrame = tk.Frame(self.notebook)
        libroDiarioFrame = tk.Frame(self.notebook)
        gastosFrame = tk.Frame(self.notebook)
        devolucionesFrame = tk.Frame(self.notebook)
        
        # Añadir los frames como pestañas al notebook
        self.notebook.add(estadoFinancieroFrame, text='💰 Estado Financiero')
        self.notebook.add(reportesFrame, text='📊 Reportes')
        self.notebook.add(libroDiarioFrame, text='📖 Libro Diario')
        self.notebook.add(gastosFrame, text='💸 Gastos')
        self.notebook.add(devolucionesFrame, text='↩️ Devoluciones')

        # Llamar a los métodos para crear los widgets de cada pestaña
        self.createEstadoFinancieroWidgets(estadoFinancieroFrame)
        self.createReportesWidgets(reportesFrame)
        self.createLibroDiarioWidgets(libroDiarioFrame)
        self.createGastosWidgets(gastosFrame)
        self.createDevolucionesWidgets(devolucionesFrame)

        tk.Button(self, text="Cerrar Ventana", command=self.onClose).pack(pady=10)
        
        # Configuración inicial si se pasa un diccionario de configuración
        if config:
            initial_tab = config.get('initial_tab', 0)
            self.notebook.select(initial_tab)
            if 'reporte' in config:
                self.reporteVar.set(config['reporte'])
            if 'periodo' in config:
                self.periodoVar.set(config['periodo'])
                self.periodoLibro.set(config['periodo'])
            self.updateView()
            self.refreshLibroDiario()
    
    def onClose(self):
        """Al cerrar, actualiza las métricas del dashboard principal."""
        if hasattr(self.parent_dashboard, 'updateDashboardMetrics'):
            self.parent_dashboard.updateDashboardMetrics()
            self.parent_dashboard.updateAnalisisGraphs()
        self.destroy()

    def createEstadoFinancieroWidgets(self, parent):
        """Crea los widgets para la pestaña 'Estado Financiero'."""
        parent.columnconfigure(1, weight=1)
        
        tk.Label(parent, text="Saldo Inicial (de config.ini):", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky="w", padx=10, pady=5)
        self.saldoInicialVar = tk.StringVar()
        tk.Entry(parent, textvariable=self.saldoInicialVar, font=("Arial", 11)).grid(row=0, column=1, sticky="ew", padx=10)
        tk.Button(parent, text="Guardar Saldo", command=self.guardarSaldoInicial).grid(row=0, column=2, padx=10)
        
        ttk.Separator(parent, orient='horizontal').grid(row=1, columnspan=3, sticky='ew', pady=10)
        
        controlesFrame = tk.Frame(parent)
        controlesFrame.grid(row=2, columnspan=3, padx=10, pady=5)
        tk.Label(controlesFrame, text="Calcular para el período:").pack(side="left")
        self.periodoEstado = tk.StringVar(value='mes')
        ttk.Radiobutton(controlesFrame, text="Día", variable=self.periodoEstado, value='dia', command=self.actualizarEstadoFinanciero).pack(side="left")
        ttk.Radiobutton(controlesFrame, text="Semana", variable=self.periodoEstado, value='semana', command=self.actualizarEstadoFinanciero).pack(side="left")
        ttk.Radiobutton(controlesFrame, text="Mes", variable=self.periodoEstado, value='mes', command=self.actualizarEstadoFinanciero).pack(side="left")
        
        self.textEstado = tk.Text(parent, height=10, width=50, font=("Courier", 12), relief="solid", bd=1, state='disabled')
        self.textEstado.grid(row=3, columnspan=3, padx=10, pady=10, sticky="ew")
        
        self.cargarSaldoInicial()
        self.actualizarEstadoFinanciero()

    def cargarSaldoInicial(self):
        """Carga el saldo inicial desde el archivo de configuración."""
        config = configparser.ConfigParser()
        config.read(CONFIG_FILE)
        balance = config.getfloat('Finance', 'starting_balance', fallback=0.0)
        self.saldoInicialVar.set(f"{balance:.2f}")

    def guardarSaldoInicial(self):
        """Guarda el valor del campo de texto como nuevo saldo inicial en el archivo de configuración."""
        try:
            nuevo_saldo = float(self.saldoInicialVar.get())
            config = configparser.ConfigParser()
            config.read(CONFIG_FILE)
            if not config.has_section('Finance'):
                config.add_section('Finance')
            config.set('Finance', 'starting_balance', str(nuevo_saldo))
            with open(CONFIG_FILE, 'w') as f:
                config.write(f)
            messagebox.showinfo("Éxito", "Saldo inicial guardado.", parent=self)
            self.actualizarEstadoFinanciero()
        except ValueError:
            messagebox.showerror("Error", "Por favor, ingrese un número válido.", parent=self)

    def actualizarEstadoFinanciero(self):
        """Calcula y muestra el estado financiero (balance de caja) para el período seleccionado."""
        periodo = self.periodoEstado.get()
        try:
            saldo_inicial = float(self.saldoInicialVar.get())
        except ValueError: saldo_inicial = 0.0
        
        with self.db.connect() as conn:
            reporte_ganancias = Venta.getReporteGanancias(conn, periodo)
        
        # Cálculo del balance
        ingresos_netos_periodo = reporte_ganancias.get('ingresosBrutos', 0) - reporte_ganancias.get('totalDescuentos', 0)
        gastos = reporte_ganancias.get('totalGastos', 0)
        devoluciones = reporte_ganancias.get('totalDevoluciones', 0)
        saldo_final = saldo_inicial + ingresos_netos_periodo - gastos - devoluciones
        
        # Formateo del texto para mostrarlo
        texto = f"Cálculo para el Período: {periodo.upper()}\n"
        texto += "----------------------------------------\n"
        texto += f"{'Saldo Inicial en Caja:':<28} ${saldo_inicial:>12.2f}\n"
        texto += f"{'(+) Ventas Netas del Periodo:':<28} ${ingresos_netos_periodo:>12.2f}\n"
        texto += f"{'(-) Devoluciones en Efectivo:':<28} -${devoluciones:>11.2f}\n"
        texto += f"{'(-) Otros Gastos Registrados:':<28} -${gastos:>11.2f}\n"
        texto += "========================================\n"
        texto += f"{'SALDO FINAL ESTIMADO EN CAJA:':<28} ${saldo_final:>12.2f}\n"
        
        self.textEstado.config(state='normal')
        self.textEstado.delete("1.0", tk.END)
        self.textEstado.insert("1.0", texto)
        self.textEstado.config(state='disabled')

    def createReportesWidgets(self, parent):
        """Crea los widgets para la pestaña 'Reportes'."""
        topFrame = tk.Frame(parent, pady=5)
        topFrame.pack(fill="x", padx=10)
        
        self.reporteVar = tk.StringVar(value='ventas')
        tk.Radiobutton(topFrame, text="Ventas", variable=self.reporteVar, value='ventas', command=self.updateView).pack(side="left")
        tk.Radiobutton(topFrame, text="Ganancias", variable=self.reporteVar, value='ganancias', command=self.updateView).pack(side="left")

        self.periodoVar = tk.StringVar(value='dia')
        tk.Radiobutton(topFrame, text="Día", variable=self.periodoVar, value='dia', command=self.updateView).pack(side="left", padx=(20,0))
        ttk.Radiobutton(topFrame, text="Semana", variable=self.periodoVar, value='semana', command=self.updateView).pack(side="left")
        ttk.Radiobutton(topFrame, text="Mes", variable=self.periodoVar, value='mes', command=self.updateView).pack(side="left")

        self.textReporte = tk.Text(parent, height=20, width=80, font=("Courier", 10), relief="solid", bd=1)
        self.textReporte.pack(pady=10, padx=10, fill="both", expand=True)
        self.updateView()
        
    def createLibroDiarioWidgets(self, parent):
        """Crea los widgets para la pestaña 'Libro Diario'."""
        controlesFrame = tk.Frame(parent, pady=5)
        controlesFrame.pack(fill="x", padx=10)
        
        tk.Label(controlesFrame, text="Ver período:").pack(side="left")
        self.periodoLibro = tk.StringVar(value='dia')
        
        ttk.Radiobutton(controlesFrame, text="Día", variable=self.periodoLibro, value='dia', command=self.refreshLibroDiario).pack(side="left", padx=5)
        ttk.Radiobutton(controlesFrame, text="Semana", variable=self.periodoLibro, value='semana', command=self.refreshLibroDiario).pack(side="left")
        ttk.Radiobutton(controlesFrame, text="Mes", variable=self.periodoLibro, value='mes', command=self.refreshLibroDiario).pack(side="left")
        tk.Label(controlesFrame, text=" (Doble clic en una venta para reimprimir y abrir ticket)").pack(side="left", padx=20)

        tree_frame = tk.Frame(parent)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("Fecha/Hora", "Descripción", "Monto", "Tipo", "ID")
        self.libroTree = ttk.Treeview(tree_frame, columns=cols, show='headings')
        self.libroTree['displaycolumns'] = ("Fecha/Hora", "Descripción", "Monto")
        for col in ("Fecha/Hora", "Descripción", "Monto"): self.libroTree.heading(col, text=col)
        self.libroTree.column("Fecha/Hora", width=160); self.libroTree.column("Descripción", width=400); self.libroTree.column("Monto", width=120, anchor="e")
        self.libroTree.tag_configure('ingreso', foreground='green'); self.libroTree.tag_configure('egreso', foreground='red')
        self.libroTree.pack(side="left", fill="both", expand=True)
        self.libroTree.bind("<Double-1>", self.reimprimirTicket)

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.libroTree.yview)
        scrollbar.pack(side="right", fill="y")
        self.libroTree.configure(yscrollcommand=scrollbar.set)
        self.refreshLibroDiario()
    
    def reimprimirTicket(self, event):
        """Se activa con doble clic en una venta del libro diario para reimprimir su ticket."""
        if not self.libroTree.focus(): return
        item_id = self.libroTree.focus()
        item_data = self.libroTree.item(item_id)
        
        full_values = item_data['values']
        tipo_transaccion = full_values[3]
        id_transaccion = full_values[4]

        if tipo_transaccion == 'venta':
            try:
                with self.db.connect() as conn:
                    ventaData = Venta.getById(conn, id_transaccion)
                if not ventaData: 
                    messagebox.showerror("Error", "No se encontraron los datos de la venta.", parent=self)
                    return
                
                # Prepara los datos necesarios para la función de generar PDF
                carrito_reimpresion = [{'nombre': d['nombre'], 'cantidad': d['cantidad'], 'subtotal': d['subtotal']} for d in ventaData['detalles']]
                pagoInfo_reimpresion = {'metodo': ventaData.get('metodoPago', 'N/A'), 'efectivo': 0, 'cambio': 0}

                ticketFile = generarTicketPdf(carrito_reimpresion, ventaData['totalVenta'], ventaData['idVenta'], pagoInfo_reimpresion)
                messagebox.showinfo("Ticket Generado", f"Se ha reimpreso el ticket:\n{ticketFile}", parent=self)

                try:
                    # Intenta abrir el PDF generado automáticamente
                    filepath_abs = os.path.abspath(ticketFile)
                    webbrowser.open(f"file:///{filepath_abs}")
                except Exception as e:
                    messagebox.showerror("Error al abrir PDF", f"No se pudo abrir el archivo PDF automáticamente:\n{e}", parent=self)

            except Exception as e:
                messagebox.showerror("Error al reimprimir", f"No se pudo generar el ticket.\n{e}", parent=self)

    def createGastosWidgets(self, parent):
        """Crea los widgets para la pestaña 'Gastos'."""
        registroFrame = tk.LabelFrame(parent, text="Registrar Nuevo Gasto", padx=10, pady=10)
        registroFrame.pack(fill="x", padx=10, pady=10)
        tk.Label(registroFrame, text="Descripción:").grid(row=0, column=0, sticky="w")
        self.gastoDescVar = tk.StringVar()
        tk.Entry(registroFrame, textvariable=self.gastoDescVar, width=40).grid(row=0, column=1, padx=5, pady=2)
        tk.Label(registroFrame, text="Monto: $").grid(row=1, column=0, sticky="w")
        self.gastoMontoVar = tk.DoubleVar()
        tk.Entry(registroFrame, textvariable=self.gastoMontoVar, width=15).grid(row=1, column=1, sticky="w", padx=5, pady=2)
        tk.Button(registroFrame, text="Registrar Gasto", command=self.registrarGasto, bg="#2ECC71", fg="white").grid(row=0, column=2, rowspan=2, padx=10, pady=2, ipady=5)
        
        listaFrame = tk.LabelFrame(parent, text="Gastos Registrados Hoy", padx=10, pady=10)
        listaFrame.pack(fill="both", expand=True, padx=10, pady=5)

        treeContainer = tk.Frame(listaFrame)
        treeContainer.pack(fill="both", expand=True)

        cols = ("Fecha", "Descripción", "Monto")
        self.gastosTree = ttk.Treeview(treeContainer, columns=cols, show='headings')
        self.gastosTree['displaycolumns'] = ("Fecha", "Descripción", "Monto") 
        for col in cols: self.gastosTree.heading(col, text=col)
        self.gastosTree.column("Fecha", width=150)
        self.gastosTree.column("Descripción", width=300)
        self.gastosTree.column("Monto", width=100, anchor="e")
        
        scrollbar = ttk.Scrollbar(treeContainer, orient="vertical", command=self.gastosTree.yview)
        self.gastosTree.configure(yscrollcommand=scrollbar.set)
        
        self.gastosTree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        tk.Button(listaFrame, text="Eliminar Gasto Seleccionado", command=self.deleteGasto, bg="#E74C3C", fg="white").pack(pady=(10,0))
        
        self.refreshGastosHoy()

    def createDevolucionesWidgets(self, parent):
        """Crea los widgets para la pestaña 'Devoluciones'."""
        tk.Label(parent, text="Buscar Ticket para Devolución", font=("Arial", 12)).pack(pady=10)
        entryFrame = tk.Frame(parent)
        entryFrame.pack(pady=5)
        tk.Label(entryFrame, text="ID del Ticket:").pack(side="left")
        self.ticketIdEntry = tk.Entry(entryFrame)
        self.ticketIdEntry.pack(side="left", padx=5)
        tk.Button(entryFrame, text="Buscar", command=self.searchSaleForReturn).pack(side="left")

    def registrarGasto(self):
        """Registra un nuevo gasto en la base de datos y actualiza las vistas."""
        descripcion, monto = self.gastoDescVar.get(), self.gastoMontoVar.get()
        if not descripcion or monto <= 0:
            messagebox.showerror("Datos inválidos", "Ingrese una descripción y un monto mayor a cero.", parent=self)
            return
        try:
            with self.db.connect() as conn: Gasto.create(conn, descripcion, monto)
            messagebox.showinfo("Éxito", "Gasto registrado.", parent=self)
            self.gastoDescVar.set(""); self.gastoMontoVar.set(0.0)
            # Actualiza todas las vistas relevantes
            self.refreshGastosHoy()
            self.updateView()
            self.actualizarEstadoFinanciero()
        except Exception as e: messagebox.showerror("Error", f"No se pudo registrar el gasto:\n{e}", parent=self)

    def deleteGasto(self):
        """Elimina el gasto seleccionado del Treeview y de la base de datos."""
        selected_item = self.gastosTree.focus()
        if not selected_item:
            messagebox.showwarning("Selección Requerida", "Por favor, seleccione un gasto de la lista para eliminar.", parent=self)
            return
        
        item_values = self.gastosTree.item(selected_item, 'values')
        gastoId = self.gastosTree.item(selected_item, 'text') # El ID se guarda en el 'text' del item
        
        confirm = messagebox.askyesno("Confirmar Eliminación", f"¿Está seguro de que desea eliminar el gasto '{item_values[1]}' por un monto de {item_values[2]}?", parent=self)
        if confirm:
            try:
                with self.db.connect() as conn:
                    Gasto.delete(conn, gastoId)
                messagebox.showinfo("Éxito", "Gasto eliminado correctamente.", parent=self)
                self.refreshGastosHoy()
                self.actualizarEstadoFinanciero()
                self.refreshLibroDiario()
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo eliminar el gasto: {e}", parent=self)

    def refreshGastosHoy(self):
        """Actualiza la lista de gastos mostrados en el Treeview."""
        for i in self.gastosTree.get_children(): self.gastosTree.delete(i)
        fechaHoy = datetime.now().strftime('%Y-%m-%d')
        with self.db.connect() as conn: 
            gastosHoy = Gasto.getByDate(conn, fechaHoy)
        for idGasto, fecha, desc, monto in gastosHoy:
            self.gastosTree.insert("", "end", text=idGasto, values=(fecha, desc, f"${monto:.2f}"))
        
    def refreshLibroDiario(self):
        """Actualiza la lista de transacciones en el libro diario."""
        for i in self.libroTree.get_children(): self.libroTree.delete(i)
        periodo = self.periodoLibro.get()
        with self.db.connect() as conn: historial = Venta.getLibroDiario(conn, periodo)
        for fecha, desc, monto, tipo, id_transaccion in historial:
            tag = 'ingreso' if monto >= 0 else 'egreso' # Asigna un tag para colorear la fila
            self.libroTree.insert("", "end", values=(fecha, desc, f"${monto:,.2f}", tipo, id_transaccion), tags=(tag,))

    def searchSaleForReturn(self):
        """Busca una venta por su ID y, si la encuentra, abre la ventana de devolución."""
        try:
            ventaId = int(self.ticketIdEntry.get())
            with self.db.connect() as conn: ventaData = Venta.getById(conn, ventaId)
            if ventaData: 
                DevolucionWindow(self, self.db, ventaData)
            else: 
                messagebox.showerror("Error", f"No se encontró la venta con ID: {ventaId}", parent=self)
        except (ValueError, TypeError): 
            messagebox.showerror("Error", "ID de ticket inválido.", parent=self)

    def updateView(self):
        """Llama al método correcto para mostrar el reporte seleccionado (ventas o ganancias)."""
        if self.reporteVar.get() == 'ventas': 
            self.showSalesReport()
        else: 
            self.showProfitReport()

    def showSalesReport(self): 
        """Genera y muestra el reporte de ventas en el campo de texto."""
        self.textReporte.delete("1.0", tk.END)
        try:
            with self.db.connect() as conn: reporte = Venta.getReporteVentas(conn, self.periodoVar.get())
            titulo = f"REPORTE DE VENTAS ({self.periodoVar.get().upper()})"
            texto = f"{titulo}\n{'='*len(titulo)}\n\n"
            texto += f"Ventas Brutas:      ${reporte['totalBruto']:>10.2f}\n"
            texto += f"Descuentos:        -${reporte['totalDescuentos']:>10.2f}\n"
            texto += f"Devoluciones:      -${reporte['totalDevoluciones']:>10.2f}\n"
            texto += "---------------------------------\n"
            texto += f"Ventas Netas:       ${reporte['ventasNetas']:>10.2f}\n\n"
            texto += f"Número de Tickets:    {reporte['numTickets']}\n\n"
            texto += "--- Productos Más Vendidos (por Cantidad) ---\n"
            if reporte['productosMasVendidos']:
                for prod, cant in reporte['productosMasVendidos']: texto += f"- {prod:<30} | Unidades: {cant}\n"
            else: texto += "No hay datos de productos para este período.\n"
            self.textReporte.insert("1.0", texto)
        except Exception as e: messagebox.showerror("Error", f"No se pudo generar reporte: {e}", parent=self)

    def showProfitReport(self):
        """Genera y muestra el reporte de ganancias en el campo de texto."""
        self.textReporte.delete("1.0", tk.END)
        try:
            with self.db.connect() as conn:
                reporte = Venta.getReporteGanancias(conn, self.periodoVar.get())
            
            # Cálculos para desglosar la ganancia
            ingresos_netos_totales = reporte['ingresosBrutos'] - reporte['totalDescuentos']
            ingresos_netos_productos = ingresos_netos_totales - reporte.get('ingresoTotalRecargas', 0)
            ganancia_de_productos = ingresos_netos_productos - reporte['costosTotales']
            ganancia_de_recargas = reporte['gananciaRecargas']
            ganancia_operativa = ganancia_de_productos + ganancia_de_recargas
            ganancia_neta_estimada = ingresos_netos_totales - reporte['totalDevoluciones'] - reporte['totalGastos']
            periodo_str = self.periodoVar.get().upper()
            
            # Formateo del texto
            titulo = f"REPORTE DE GANANCIAS ({periodo_str})"
            texto = f"{titulo}\n{'='*len(titulo)}\n\n"
            texto += "--- 1. DESGLOSE DE INGRESOS ---\n"
            texto += f"{'Ingresos Netos (Productos):':<28} ${ingresos_netos_productos:>12.2f}\n"
            texto += f"{'Ingresos Netos (Recargas):':<28} ${reporte.get('ingresoTotalRecargas', 0):>12.2f}\n"
            texto += "----------------------------------------\n"
            texto += f"{'Ingresos Netos Totales:':<28} ${ingresos_netos_totales:>12.2f}\n\n"
            texto += "--- 2. GANANCIA OPERATIVA ---\n"
            texto += f"{'Ganancia por Productos:':<28} ${ganancia_de_productos:>12.2f}\n"
            texto += f"  (Ingresos: ${ingresos_netos_productos:.2f} - Costo: ${reporte['costosTotales']:.2f})\n"
            texto += f"{'(+) Ganancia Pura (Recargas):':<28} ${ganancia_de_recargas:>12.2f}\n"
            texto += "----------------------------------------\n"
            texto += f"{'Ganancia Operativa Total:':<28} ${ganancia_operativa:>12.2f}\n\n"
            texto += "--- 3. GANANCIA NETA FINAL ---\n"
            texto += f"{'Ganancia Operativa:':<28} ${ingresos_netos_totales:>12.2f}\n"
            texto += f"{'(-) Devoluciones en Efectivo:':<28} -${reporte['totalDevoluciones']:>11.2f}\n"
            texto += f"{'(-) Otros Gastos Registrados:':<28} -${reporte['totalGastos']:>11.2f}\n"
            texto += "========================================\n"
            texto += f"{'GANANCIA NETA ESTIMADA:':<28} ${ganancia_neta_estimada:>12.2f}\n"
            self.textReporte.insert("1.0", texto)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el reporte de ganancias:\n{e}", parent=self)

class AdminInventarioWindow(tk.Toplevel):
    """
    Ventana completa para la gestión de inventario: agregar, editar, eliminar,
    reabastecer, buscar, filtrar y exportar productos.
    """
    def __init__(self, parent, db_instance, *args):
        super().__init__(parent)
        self.db = db_instance
        self.title("Administración de Inventario")
        self.geometry("1200x600")
        self.protocol("WM_DELETE_WINDOW", self.destroy)

        # --- Frame de Controles (Filtros y Búsqueda) ---
        controls_frame = tk.Frame(self, pady=5)
        controls_frame.pack(fill="x", padx=10)
        tk.Label(controls_frame, text="Filtrar por Categoría:").pack(side="left")
        self.categoriaFilter = ttk.Combobox(controls_frame, state="readonly", width=25)
        self.categoriaFilter.pack(side="left", padx=(5, 10))
        self.categoriaFilter.bind("<<ComboboxSelected>>", self.refreshList)
        tk.Button(controls_frame, text="Limpiar Filtro", command=self.clearFilter).pack(side="left")
        tk.Label(controls_frame, text="Buscar Producto:").pack(side="left", padx=(20, 5))
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self.onSearch)
        search_entry = tk.Entry(controls_frame, textvariable=self.search_var, width=40)
        search_entry.pack(side="left", fill="x", expand=True)

        # --- Treeview para mostrar el inventario ---
        tree_frame = tk.Frame(self)
        tree_frame.pack(pady=10, padx=10, fill="both", expand=True)
        cols = ("ID", "Código", "Nombre", "Categoría", "Precio", "Costo", "Stock")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show='headings')
        for col in cols: self.tree.heading(col, text=col)
        self.tree.column("ID", width=40, anchor="center"); self.tree.column("Código", width=130); self.tree.column("Nombre", width=250); self.tree.column("Categoría", width=120); self.tree.column("Precio", width=80, anchor="e"); self.tree.column("Costo", width=80, anchor="e"); self.tree.column("Stock", width=60, anchor="center")
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # --- Frame de Botones de Acción ---
        action_frame = tk.Frame(self, padx=10, pady=10)
        action_frame.pack(fill="x")
        tk.Button(action_frame, text="Agregar", command=self.addProduct, bg="#2ECC71", fg="white").pack(side="left", padx=5)
        tk.Button(action_frame, text="Editar", command=self.editProduct, bg="#F1C40F").pack(side="left", padx=5)
        tk.Button(action_frame, text="Reabastecer", command=self.restockProduct, bg="#16A085", fg="white").pack(side="left", padx=5)
        tk.Button(action_frame, text="Eliminar", command=self.deleteProduct, bg="#E74C3C", fg="white").pack(side="left", padx=5)
        tk.Button(action_frame, text="Importación Rápida", command=self.abrirDialogoImportacion, bg="#007BFF", fg="white").pack(side="left", padx=5)
        tk.Button(action_frame, text="Exportar a CSV", command=self.exportInventoryToCsv).pack(side="left", padx=5)
        tk.Button(action_frame, text="Exportar a Excel", command=self.exportInventoryToXlsx).pack(side="left", padx=5)
        tk.Button(action_frame, text="Cerrar", command=self.destroy).pack(side="right", padx=5)
        
        # Configuración de tag para colorear filas con bajo stock
        self.tree.tag_configure('low_stock', background='#E74C3C', foreground='white')
        
        self.loadCategories()
        self.refreshList()

    def abrirDialogoImportacion(self):
        """Abre el diálogo para importación masiva de productos."""
        dialogo = DialogoImportacionTexto(self, self.db)
        self.wait_window(dialogo)
        if dialogo.importacionExitosa:
            self.refreshList()

    def onSearch(self, *args):
        """Se activa al escribir en el campo de búsqueda para filtrar la lista."""
        term = self.search_var.get()
        if len(term) > 1:
            with self.db.connect() as conn:
                results = Producto.searchInventory(conn, term)
            self.refreshList(lista_productos=results)
        elif not term: # Si se borra la búsqueda, muestra toda la lista de nuevo
            self.refreshList()

    def loadCategories(self):
        """Carga las categorías de la base de datos y las pone en el Combobox de filtro."""
        with self.db.connect() as conn:
            self.categoriasData = Categoria.getAll(conn)
        self.categoriasMap = {nombre: catId for catId, nombre in self.categoriasData}
        self.categoriaFilter['values'] = ['Todas'] + sorted(list(self.categoriasMap.keys()))
        self.categoriaFilter.set('Todas')

    def clearFilter(self):
        """Limpia todos los filtros y refresca la lista."""
        self.categoriaFilter.set('Todas')
        self.search_var.set("")
        self.refreshList()

    def refreshList(self, event=None, lista_productos=None):
        """Actualiza el Treeview con la lista de productos, aplicando filtros si es necesario."""
        for i in self.tree.get_children(): self.tree.delete(i)
        
        productList = []
        if lista_productos is not None: # Si se pasa una lista de productos (desde la búsqueda)
            productList = lista_productos
        else: # Si no, se obtienen los productos según el filtro de categoría
            selectedCategory = self.categoriaFilter.get()
            categoriaId = self.categoriasMap.get(selectedCategory) if selectedCategory != 'Todas' else None
            try:
                with self.db.connect() as conn:
                    productList = Producto.getAll(conn, categoriaId)
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo cargar el inventario: {e}", parent=self)

        for prod in productList:
            # Asigna el tag 'low_stock' si el stock es <= 5 y no es una recarga
            tags = ('low_stock',) if len(prod) > 6 and prod[6] <= 5 and prod[2] != "Recarga Celular" else ()
            self.tree.insert("", "end", values=prod, tags=tags)

    def exportInventoryToCsv(self):
        """Exporta el inventario completo a un archivo CSV."""
        filepath = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("Archivos CSV", "*.csv")])
        if not filepath: return
        try:
            with self.db.connect() as conn:
                all_products = Producto.getAll(conn)
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["ID", "Codigo de Barras", "Nombre", "Categoria", "Precio Venta", "Costo Compra", "Stock"])
                writer.writerows(all_products)
            messagebox.showinfo("Éxito", f"Inventario exportado a\n{filepath}", parent=self)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar el archivo: {e}", parent=self)

    def exportInventoryToXlsx(self):
        """Exporta el inventario completo a un archivo de Excel."""
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos de Excel", "*.xlsx")])
        if not filepath: return
        try:
            with self.db.connect() as conn:
                all_products = Producto.getAll(conn)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventario"
            headers = ["ID", "Codigo de Barras", "Nombre", "Categoria", "Precio Venta", "Costo Compra", "Stock"]
            ws.append(headers)
            for product_tuple in all_products:
                ws.append(product_tuple)
            wb.save(filepath)
            messagebox.showinfo("Éxito", f"Inventario exportado a\n{filepath}", parent=self)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar el archivo Excel: {e}", parent=self)

    def addProduct(self):
        """Abre el diálogo de producto para crear uno nuevo."""
        self.openProductDialog()

    def editProduct(self):
        """Abre el diálogo de producto con los datos del producto seleccionado para editarlo."""
        if not self.tree.focus():
            messagebox.showwarning("Selección Requerida", "Por favor, seleccione un producto para editar.", parent=self)
            return
        productId = self.tree.item(self.tree.focus())['values'][0]
        with self.db.connect() as conn:
            productData = Producto.getById(conn, productId)
        if productData:
            self.openProductDialog(productData)

    def deleteProduct(self):
        """Elimina el producto seleccionado."""
        if not self.tree.focus():
            messagebox.showwarning("Selección Requerida", "Por favor, seleccione un producto para eliminar.", parent=self)
            return
        values = self.tree.item(self.tree.focus())['values']
        # Protección para no eliminar productos especiales
        if values[2] == "Recarga Celular":
            messagebox.showerror("Error", "El producto 'Recarga Celular' no puede ser eliminado.", parent=self)
            return
        if messagebox.askyesno("Confirmar", f"¿Eliminar el producto '{values[2]}'?"):
            with self.db.connect() as conn:
                Producto.delete(conn, values[0])
            self.refreshList()

    def restockProduct(self):
        """Abre un diálogo para agregar stock a un producto seleccionado."""
        if not self.tree.focus():
            messagebox.showwarning("Selección Requerida", "Por favor, seleccione un producto para reabastecer.", parent=self)
            return
        values = self.tree.item(self.tree.focus())['values']
        productId = values[0]
        productName = values[2]
        qty = simpledialog.askinteger("Reabastecer", f"Unidades a agregar al stock de '{productName}':", parent=self, minvalue=1)
        if qty:
            with self.db.connect() as conn:
                Producto.updateStock(conn, productId, qty)
            messagebox.showinfo("Reabastecer", f"{qty} unidades de '{productName}' agregadas al stock.", parent=self)
            self.refreshList()

    def openProductDialog(self, producto=None):
        """
        Diálogo para agregar o editar un producto.
        Si 'producto' es None, es para agregar. Si no, es para editar.
        """
        dialog = tk.Toplevel(self)
        dialog.title("Editar Producto" if producto else "Agregar Producto")
        dialog.grab_set()
        
        # Campos del formulario
        fields = {"C. Barras:": tk.StringVar(), "Nombre:": tk.StringVar(), "Categoría:": None, "Precio:": tk.DoubleVar(), "Costo:": tk.DoubleVar(), "Stock:": tk.IntVar()}
        
        # Si se está editando, se llenan los campos con los datos del producto
        if producto:
            fields["C. Barras:"].set(producto['codigoBarras'])
            fields["Nombre:"].set(producto['nombre'])
            fields["Precio:"].set(producto['precioVenta'])
            fields["Costo:"].set(producto['costoCompra'])
            fields["Stock:"].set(producto['stock'])
            
        # Desactiva campos para productos especiales que no deben ser modificados
        is_recharge_product = producto and producto['nombre'] == "Recarga Celular"
        
        for i, label_text in enumerate(fields):
            tk.Label(dialog, text=label_text).grid(row=i, column=0, padx=5, pady=2, sticky="w")
            if label_text == "Categoría:":
                catFrame = tk.Frame(dialog)
                catFrame.grid(row=i, column=1, sticky="ew", pady=2)
                categoriaCombo = ttk.Combobox(catFrame, state="readonly")
                categoriaCombo.pack(side="left", expand=True, fill="x")
                fields[label_text] = categoriaCombo
                if is_recharge_product:
                    categoriaCombo.config(state='disabled')
            else:
                entry_widget = tk.Entry(dialog, textvariable=fields[label_text])
                entry_widget.grid(row=i, column=1, padx=5, pady=2, sticky="ew")
                if is_recharge_product and label_text in ["C. Barras:", "Nombre:", "Precio:", "Costo:"]:
                    entry_widget.config(state='disabled')
                    
        # Carga de categorías en el combobox
        with self.db.connect() as conn:
            allCategorias = {name: cid for cid, name in Categoria.getAll(conn)}
        fields["Categoría:"]['values'] = sorted(list(allCategorias.keys()))
        if producto and producto.get('idCategoria'):
            catName = [name for name, cid in allCategorias.items() if cid == producto['idCategoria']]
            if catName: fields["Categoría:"].set(catName[0])
            
        def addNewCategory():
            """Función para agregar una nueva categoría desde el diálogo de producto."""
            newCatName = simpledialog.askstring("Nueva Categoría", "Nombre de la nueva categoría:", parent=dialog)
            if newCatName:
                try:
                    with self.db.connect() as conn:
                        Categoria.create(conn, newCatName)
                    self.loadCategories() # Recarga las categorías en la ventana principal
                    dialog.destroy() # Cierra y reabre el diálogo para que aparezca la nueva categoría
                    self.openProductDialog(producto)
                except Exception as e:
                    messagebox.showerror("Error", str(e), parent=dialog)
        
        btn_add_cat = tk.Button(catFrame, text="+", command=addNewCategory)
        btn_add_cat.pack(side="left", padx=2)
        if is_recharge_product:
            btn_add_cat.config(state='disabled')
            
        def save():
            """Guarda los datos del producto (nuevo o editado)."""
            try:
                # Lógica especial para actualizar solo el stock de recargas
                if is_recharge_product:
                    new_stock = fields["Stock:"].get()
                    if new_stock < 0: raise ValueError("El stock no puede ser negativo.")
                    with self.db.connect() as conn:
                        current_stock = Producto.getById(conn, producto['idProducto'])['stock']
                        stock_change = new_stock - current_stock
                        Producto.updateStock(conn, producto['idProducto'], stock_change)
                else: # Lógica para productos normales
                    catId = allCategorias.get(fields["Categoría:"].get())
                    with self.db.connect() as conn:
                        if producto: # Actualizar
                            Producto.update(conn, producto['idProducto'], fields["C. Barras:"].get(), fields["Nombre:"].get(), fields["Precio:"].get(), fields["Costo:"].get(), fields["Stock:"].get(), catId)
                        else: # Crear
                            Producto.create(conn, fields["C. Barras:"].get(), fields["Nombre:"].get(), fields["Precio:"].get(), fields["Costo:"].get(), fields["Stock:"].get(), catId)
                self.refreshList()
                dialog.destroy()
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=dialog)
                
        tk.Button(dialog, text="Guardar", command=save).grid(row=len(fields), columnspan=2, pady=10)
        dialog.bind("<Return>", lambda e: save())

class AdminUsuariosWindow(tk.Toplevel):
    """Ventana para la administración de usuarios (solo accesible para administradores)."""
    def __init__(self, parent, db_instance, currentAdminUsername):
        super().__init__(parent)
        self.parent = parent
        self.db = db_instance
        self.currentAdminUsername = currentAdminUsername
        self.title("Administración de Usuarios")
        self.geometry("600x400")
        self.protocol("WM_DELETE_WINDOW", self.onClose)

        tree_frame = tk.Frame(self)
        tree_frame.pack(pady=10, padx=10, fill="both", expand=True)
        cols = ("ID", "Nombre de Usuario", "Rol")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show='headings', selectmode="browse")
        for col in cols: self.tree.heading(col, text=col)
        self.tree.column("ID", width=50, anchor="center")
        self.tree.pack(fill="both", expand=True)
        
        action_frame = tk.Frame(self, padx=10, pady=10)
        action_frame.pack(fill="x")
        tk.Button(action_frame, text="Agregar", command=self.addUser, bg="#2ECC71", fg="white").pack(side="left", padx=5)
        tk.Button(action_frame, text="Editar", command=self.editUser, bg="#F1C40F").pack(side="left", padx=5)
        tk.Button(action_frame, text="Eliminar", command=self.deleteUser, bg="#E74C3C", fg="white").pack(side="left", padx=5)
        tk.Button(action_frame, text="Cerrar", command=self.onClose).pack(side="right", padx=5)
        
        self.refreshList()

    def refreshList(self):
        """Actualiza la lista de usuarios en el Treeview."""
        for i in self.tree.get_children(): self.tree.delete(i)
        try:
            with self.db.connect() as conn: userList = Usuario.getAll(conn)
            for user in userList: self.tree.insert("", "end", iid=user[0], values=(user[0], user[1], user[2].capitalize()))
        except Exception as e: messagebox.showerror("Error", f"No se pudo cargar la lista de usuarios: {e}", parent=self)

    def addUser(self): 
        """Abre el diálogo de usuario para crear uno nuevo."""
        self.openUserDialog()

    def editUser(self):
        """Abre el diálogo de usuario con los datos del usuario seleccionado para editarlo."""
        if not self.tree.focus(): 
            messagebox.showwarning("Selección Requerida", "Seleccione un usuario para editar.", parent=self)
            return
        userId = int(self.tree.focus())
        with self.db.connect() as conn:
            allUsers = Usuario.getAll(conn)
            userData = next((user for user in allUsers if user[0] == userId), None)
        if userData: self.openUserDialog(userData)

    def deleteUser(self):
        """Elimina el usuario seleccionado."""
        if not self.tree.focus(): 
            messagebox.showwarning("Selección Requerida", "Seleccione un usuario para eliminar.", parent=self)
            return
        userId = int(self.tree.focus())
        username = self.tree.item(self.tree.focus())['values'][1]
        
        # El administrador no se puede eliminar a sí mismo
        if username == self.currentAdminUsername: 
            messagebox.showerror("Acción no permitida", "No puedes eliminar tu propio usuario.", parent=self)
            return
            
        if messagebox.askyesno("Confirmar Eliminación", f"¿Está seguro de que desea eliminar al usuario '{username}'?", parent=self):
            try:
                with self.db.connect() as conn: Usuario.delete(conn, userId)
                messagebox.showinfo("Éxito", "Usuario eliminado correctamente.", parent=self)
                self.refreshList()
            except Exception as e: messagebox.showerror("Error", f"No se pudo eliminar el usuario: {e}", parent=self)

    def openUserDialog(self, userData=None):
        """Diálogo para agregar o editar un usuario."""
        dialog = tk.Toplevel(self)
        dialog.title("Editar Usuario" if userData else "Agregar Usuario")
        dialog.grab_set()
        
        userId, usernameVal, roleVal = (userData[0], userData[1], userData[2].lower()) if userData else (None, "", "cajero")
        
        username = tk.StringVar(value=usernameVal)
        password = tk.StringVar()
        role = tk.StringVar(value=roleVal)
        
        tk.Label(dialog, text="Usuario:").grid(row=0,column=0,padx=5,pady=2,sticky="w")
        tk.Entry(dialog, textvariable=username).grid(row=0,column=1,padx=5,pady=2)
        
        # La etiqueta de la contraseña cambia si se está editando o creando
        pwLabelText = "Contraseña (vacío para no cambiar):" if userData else "Contraseña:"
        tk.Label(dialog, text=pwLabelText).grid(row=1,column=0,padx=5,pady=2,sticky="w")
        tk.Entry(dialog, textvariable=password, show="*").grid(row=1,column=1,padx=5,pady=2)
        
        tk.Label(dialog, text="Rol:").grid(row=2,column=0,padx=5,pady=2,sticky="w")
        ttk.Combobox(dialog, textvariable=role, values=['admin', 'cajero'], state="readonly").grid(row=2,column=1,padx=5,pady=2)
        
        def save():
            """Guarda el usuario nuevo o editado."""
            try:
                with self.db.connect() as conn:
                    if userData: # Actualizar
                        Usuario.update(conn, userId, username.get(), password.get(), role.get())
                    else: # Crear
                        Usuario.create(conn, username.get(), password.get(), role.get())
                self.refreshList()
                dialog.destroy()
            except Exception as e: 
                messagebox.showerror("Error", str(e), parent=dialog)
                
        tk.Button(dialog, text="Guardar", command=save).grid(row=3, columnspan=2, pady=10)
        dialog.bind("<Return>", lambda e: save())

    def onClose(self):
        """Al cerrar, vuelve a mostrar la ventana padre (Dashboard)."""
        self.parent.deiconify()
        self.destroy()

class DevolucionWindow(tk.Toplevel):
    """Ventana para procesar la devolución de productos de una venta específica."""
    def __init__(self, parent, db_instance, venta_data):
        super().__init__(parent)
        self.db = db_instance
        self.ventaData = venta_data
        self.title(f"Devolución - Ticket #{self.ventaData['idVenta']}")
        self.geometry("700x400")
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.itemsParaDevolver = {} # Diccionario para llevar la cuenta de los items a devolver

        infoFrame = tk.LabelFrame(self, text="Información de la Venta Original")
        infoFrame.pack(fill="x", padx=10, pady=5)
        tk.Label(infoFrame, text=f"Fecha: {self.ventaData['fecha']} | Total Pagado: ${self.ventaData['totalVenta']:.2f}").pack()
        
        tree_frame = tk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)
        cols = ("Producto", "Cant. Comprada", "Precio Unit.", "Cant. a Devolver")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        for col in cols: self.tree.heading(col, text=col)
        self.tree.pack(side="left", fill="both", expand=True)
        
        # Llena el Treeview con los productos de la venta original
        for item in self.ventaData['detalles']:
            self.tree.insert("", "end", iid=item['idProducto'], values=(item['nombre'], item['cantidad'], f"${item['precioUnitario']:.2f}", 0))
            
        tk.Button(self, text="Seleccionar Item para Devolver", command=self.selectItem).pack(pady=5)
        tk.Button(self, text="Procesar Devolución", command=self.processReturn, bg="#2ECC71", fg="white").pack(pady=5)
        tk.Button(self, text="Cerrar", command=self.destroy).pack(pady=5)

    def selectItem(self):
        """Permite al usuario seleccionar un item y especificar la cantidad a devolver."""
        if not self.tree.focus(): return
        productoId = int(self.tree.focus())
        itemValues = self.tree.item(productoId, "values")
        
        # No se pueden devolver recargas
        if itemValues[0].startswith("Recarga Celular"):
            messagebox.showwarning("Advertencia", "No se pueden devolver recargas celulares.", parent=self)
            return
            
        cantADevolver = simpledialog.askinteger("Cantidad", f"¿Unidades de '{itemValues[0]}' a devolver?", parent=self, minvalue=0, maxvalue=int(itemValues[1]))
        if cantADevolver is not None:
            nuevosValores = list(itemValues)
            nuevosValores[3] = cantADevolver # Actualiza la columna 'Cant. a Devolver'
            self.tree.item(productoId, values=nuevosValores)
            self.itemsParaDevolver[productoId] = cantADevolver # Guarda la cantidad en el diccionario

    def processReturn(self):
        """Procesa la devolución, registrándola en la BD y actualizando el stock."""
        if not any(self.itemsParaDevolver.values()): 
            messagebox.showerror("Error", "No ha seleccionado ninguna cantidad para devolver.", parent=self)
            return
            
        itemsFinales, montoTotal = [], 0
        # Prepara la lista de items a devolver y calcula el monto total
        for prodId, cantidad in self.itemsParaDevolver.items():
            if cantidad > 0:
                detalle = next(d for d in self.ventaData['detalles'] if d['idProducto'] == prodId)
                montoDevuelto = detalle['precioUnitario'] * cantidad
                montoTotal += montoDevuelto
                itemsFinales.append({"idProducto": prodId, "nombreProducto": detalle['nombre'], "cantidad": cantidad, "montoDevuelto": montoDevuelto})
                
        if messagebox.askyesno("Confirmar", f"Monto a reembolsar: ${montoTotal:.2f}.\n¿Continuar?", parent=self):
            with self.db.connect() as conn: 
                Devolucion.create(conn, self.ventaData['idVenta'], itemsFinales)
            messagebox.showinfo("Éxito", "Devolución procesada.", parent=self)
            self.destroy()

class HerramientasWindow(tk.Toplevel):
    """Proporciona herramientas críticas como la creación y restauración de copias de seguridad de la base de datos."""
    def __init__(self, parent, db_instance, *args):
        super().__init__(parent)
        self.db = db_instance
        self.rootApp = parent.rootApp
        self.title("Herramientas Administrativas")
        self.geometry("400x300")
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        tk.Label(self, text="Copia de Seguridad y Restauración", font=("Arial", 14, "bold")).pack(pady=20)
        tk.Button(self, text="Crear Copia de Seguridad Ahora", command=self.crearCopiaSeguridad, width=30, height=2).pack(pady=10)
        tk.Button(self, text="Restaurar desde Copia", command=self.restaurarCopiaSeguridad, width=30, height=2, bg="#c0392b", fg="white").pack(pady=10)
        tk.Button(self, text="Cerrar", command=self.destroy).pack(pady=10)

    def crearCopiaSeguridad(self):
        """Crea una copia del archivo de la base de datos con un timestamp."""
        backup_dir = "backups"
        if not os.path.exists(backup_dir): os.makedirs(backup_dir)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_path = os.path.join(backup_dir, f"backup-{timestamp}.db")
        try:
            shutil.copyfile(self.db.dbPath, backup_path)
            messagebox.showinfo("Éxito", f"Copia de seguridad creada con éxito en:\n{backup_path}", parent=self)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear la copia de seguridad:\n{e}", parent=self)

    def restaurarCopiaSeguridad(self):
        """Reemplaza la base de datos actual con un archivo de copia de seguridad seleccionado."""
        advertencia = "¡ADVERTENCIA!\n\nEsto reemplazará TODOS los datos actuales con los de la copia de seguridad.\n\nLa aplicación se cerrará después de restaurar. Deberá volver a abrirla.\n\n¿Está seguro de que desea continuar?"
        if not messagebox.askyesno("Confirmación Crítica", advertencia, icon='warning', parent=self): return
        
        filepath = filedialog.askopenfilename(title="Seleccione una copia de seguridad (.db)", filetypes=[("Archivos de Base de Datos", "*.db"), ("Todos los archivos", "*.*")])
        if not filepath: return
        
        try:
            shutil.copyfile(filepath, self.db.dbPath)
            messagebox.showinfo("Restauración Completa", "La base de datos ha sido restaurada.\nLa aplicación se cerrará ahora. Por favor, vuelva a abrirla.", parent=self)
            self.rootApp.destroy() # Cierra la aplicación para que los cambios surtan efecto al reabrir
        except Exception as e:
            messagebox.showerror("Error de Restauración", f"No se pudo restaurar la base de datos:\n{e}", parent=self)

class DialogoImportacionTexto(tk.Toplevel):
    """Permite la importación masiva de productos pegando texto con formato CSV."""
    def __init__(self, parent, db_instance):
        super().__init__(parent)
        self.db = db_instance
        self.importacionExitosa = False
        self.title("Importación Rápida de Productos")
        self.geometry("700x500")
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.grab_set()
        
        info_label = tk.Label(self, text="Pega los datos aquí. Formato (6 columnas separadas por comas):\nCodigo,Nombre,Precio,Costo,Stock,Categoria", justify=tk.LEFT)
        info_label.pack(pady=(10, 5), padx=10, anchor="w")
        
        text_frame = tk.Frame(self, bd=1, relief="sunken")
        text_frame.pack(pady=5, padx=10, expand=True, fill="both")
        self.text_widget = tk.Text(text_frame, wrap="word", font=("Courier", 10))
        self.text_widget.pack(side="left", expand=True, fill="both")
        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=self.text_widget.yview)
        scrollbar.pack(side="right", fill="y")
        self.text_widget.config(yscrollcommand=scrollbar.set)
        
        button_frame = tk.Frame(self)
        button_frame.pack(pady=10)
        tk.Button(button_frame, text="Procesar e Importar", command=self.procesarTexto, bg="#28a745", fg="white", font=("Arial", 12)).pack(side="left", padx=10)
        tk.Button(button_frame, text="Cerrar", command=self.destroy).pack(side="left", padx=10)

    def procesarTexto(self):
        """Lee el texto, lo procesa línea por línea e intenta importar los productos."""
        texto_completo = self.text_widget.get("1.0", tk.END)
        lineas = texto_completo.strip().split('\n')
        productos_agregados, errores = 0, []
        
        with self.db.connect() as conn:
            categoriasMap = {nombre.lower(): catId for catId, nombre in Categoria.getAll(conn)}
            
        for i, linea in enumerate(lineas):
            if not linea.strip(): continue # Ignora líneas vacías
            
            valores = linea.split(',')
            try:
                if len(valores) != 6:
                    errores.append(f"Línea {i+1}: Se esperan 6 columnas.")
                    continue
                
                codigo, nombre, precio, costo, stock, nombreCategoria = [v.strip() for v in valores]
                
                catId = None
                if nombreCategoria:
                    catId = categoriasMap.get(nombreCategoria.lower())
                    # Si la categoría no existe, la crea al vuelo
                    if not catId:
                        try:
                            with self.db.connect() as conn: Categoria.create(conn, nombreCategoria)
                            # Actualiza el mapa de categorías para obtener el ID de la nueva
                            with self.db.connect() as conn:
                                categoriasMap = {nombre.lower(): catId for catId, nombre in Categoria.getAll(conn)}
                                catId = categoriasMap.get(nombreCategoria.lower())
                        except Exception as cat_e:
                            errores.append(f"Línea {i+1}: No se pudo crear categoría '{nombreCategoria}': {cat_e}")
                
                # Intenta crear el producto
                with self.db.connect() as conn:
                    Producto.create(conn, codigo, nombre, float(precio), float(costo), int(stock), catId)
                productos_agregados += 1
                
            except ValueError:
                errores.append(f"Línea {i+1}: Revisa que precio y stock sean números.")
            except Exception as e:
                errores.append(f"Línea {i+1}: {e}")
                
        # Muestra un resumen de la importación
        resumen_msg = f"{productos_agregados} productos importados."
        if errores: resumen_msg += f"\n\nErrores:\n" + "\n".join(errores[:10]) # Muestra los primeros 10 errores
        messagebox.showinfo("Resultado de la Importación", resumen_msg, parent=self)
        
        if productos_agregados > 0: self.importacionExitosa = True
        self.destroy()

class DialogoVentaDulces(tk.Toplevel):
    """Un diálogo especializado para vender productos de la categoría 'dulces' rápidamente."""
    def __init__(self, parent, db_instance, id_categoria):
        super().__init__(parent)
        self.db = db_instance
        self.title("Seleccionar Dulces")
        self.geometry("600x400")
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.grab_set()
        self.dulcesSeleccionados = {} # Almacena temporalmente los dulces seleccionados aquí

        tk.Label(self, text="Doble clic en un dulce para agregarlo").pack(pady=5)
        
        main_frame = tk.Frame(self)
        main_frame.pack(fill="both", expand=True, padx=10, pady=5)
        cols = ("Nombre", "Precio", "Stock")
        self.tree = ttk.Treeview(main_frame, columns=cols, show='headings', selectmode="browse")
        for col in cols: self.tree.heading(col, text=col)
        self.tree.column("Precio", anchor="e"); self.tree.column("Stock", anchor="center")
        self.tree.pack(side="left", fill="both", expand=True)
        self.tree.bind("<Double-1>", self.agregarDulce)
        
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        button_frame = tk.Frame(self)
        button_frame.pack(pady=10)
        tk.Button(button_frame, text="Confirmar y Agregar", command=self.confirmar).pack(side="left", padx=10)
        tk.Button(button_frame, text="Cancelar", command=self.destroy).pack(side="left", padx=10)
        
        self.cargarDulces(id_categoria)

    def cargarDulces(self, id_categoria):
        """Carga todos los productos de la categoría especificada en el Treeview."""
        with self.db.connect() as conn:
            lista_productos_tuplas = Producto.getAll(conn, categoriaId=id_categoria)
        
        self.listaProductosDict = []
        for p_tuple in lista_productos_tuplas:
            self.tree.insert("", "end", iid=p_tuple[0], values=(p_tuple[2], f"${p_tuple[4]:.2f}", p_tuple[6]))
            # También guarda los datos completos en un diccionario para fácil acceso
            self.listaProductosDict.append(Producto.getById(self.db.connect(), p_tuple[0]))

    def agregarDulce(self, event=None):
        """Se activa con doble clic para agregar un dulce a la selección temporal."""
        if not self.tree.focus(): return
        producto_id = int(self.tree.focus())
        producto_dict = next((p for p in self.listaProductosDict if p["idProducto"] == producto_id), None)
        if not producto_dict: return
        
        cantidad = simpledialog.askinteger("Cantidad", f"¿Cuántas unidades de '{producto_dict['nombre']}'?", parent=self, minvalue=1, maxvalue=producto_dict['stock'])
        if cantidad:
            self.dulcesSeleccionados[producto_dict['idProducto']] = (producto_dict, cantidad)
            messagebox.showinfo("Agregado", f"{cantidad} x {producto_dict['nombre']} en espera.", parent=self)

    def confirmar(self):
        """Cierra el diálogo. La ventana padre se encargará de leer los datos de 'dulcesSeleccionados'."""
        self.destroy()

class LowStockWindow(tk.Toplevel):
    """Muestra una lista de todos los productos con bajo nivel de existencias."""
    def __init__(self, parent, db_instance, *args):
        super().__init__(parent)
        self.db = db_instance
        self.title("Productos con Bajo Stock (<= 5 unidades)")
        self.geometry("1000x600")
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.grab_set()
        
        tree_frame = tk.Frame(self)
        tree_frame.pack(pady=10, padx=10, fill="both", expand=True)
        cols = ("ID", "Código", "Nombre", "Categoría", "Stock")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show='headings')
        for col in cols: self.tree.heading(col, text=col)
        self.tree.pack(side="left", fill="both", expand=True)

        with self.db.connect() as conn:
            low_stock_products = Producto.getLowStock(conn, limit=5)

        for prod in low_stock_products:
            self.tree.insert("", "end", values=prod, tags=('low_stock',))
        
        # Colorea las filas con el tag 'low_stock'
        self.tree.tag_configure('low_stock', background='#E74C3C', foreground='white')
        
        tk.Button(self, text="Cerrar", command=self.destroy).pack(pady=10)

# --- Punto de Entrada de la Aplicación ---
if __name__ == "__main__":
    # 1. Crea el archivo de configuración si no existe, con valores por defecto.
    if not os.path.exists(CONFIG_FILE):
        config = configparser.ConfigParser()
        config['Login'] = {'username': ''}
        config['Finance'] = {'starting_balance': '0.0'}
        with open(CONFIG_FILE, 'w') as configfile:
            config.write(configfile)

    # 2. Inicializa la conexión a la base de datos (y crea las tablas si no existen).
    db = Database()
    
    # 3. Se asegura de que existan datos iniciales básicos para el primer uso.
    with db.connect() as conn:
        Usuario.createDefaultAdminIfNeeded(conn) # Crea el usuario 'admin'
        Producto.populateInitialProducts(conn) # Crea productos base como 'Recarga Celular'
    
    # 4. Crea la ventana raíz de Tkinter pero la mantiene oculta (withdraw).
    #    Sirve como "dueña" de todas las demás ventanas.
    appRoot = tk.Tk()
    appRoot.withdraw()

    def onLoginSuccess(role, username):
        """
        Callback que se ejecuta tras un inicio de sesión exitoso.
        Abre la ventana correspondiente al rol del usuario.
        """
        if role == 'admin':
            DashboardWindow(appRoot, username, db)
        else: # rol == 'cajero'
            PuntoVentaApp(appRoot, role, username, db)

    # 5. Inicia el flujo de la aplicación mostrando la ventana de login.
    LoginWindow(appRoot, onLoginSuccess, db)
    
    # 6. Inicia el bucle principal de eventos de Tkinter. La aplicación espera aquí
    #    la interacción del usuario.
    appRoot.mainloop()